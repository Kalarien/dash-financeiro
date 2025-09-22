#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dashboard Financeiro
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# Importa o sistema unificado
from importador_matriz_unificada import ImportadorMatrizUnificada
from modulo_contas_receber_asaas import ContasReceberAsaas

# Sistema de autenticação adaptado para Streamlit Cloud
def check_authentication():
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        st.title("🔐 Acesso ao Dashboard Financeiro")
        
        with st.form("login_form"):
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.markdown("### Login")
                username = st.text_input("Usuário", placeholder="Digite seu usuário")
                password = st.text_input("Senha", type="password", placeholder="Digite sua senha")
                submitted = st.form_submit_button("Entrar", use_container_width=True)
                
                if submitted:
                    if username == "admin" and password == "CulturaBuilder852@":
                        st.session_state.authenticated = True
                        st.success("Login realizado com sucesso!")
                        st.rerun()
                    else:
                        st.error("Usuário ou senha incorretos")
        
        st.stop()

def show_user_info():
    with st.sidebar:
        st.markdown("---")
        st.markdown("👤 **Usuário:** admin")
        if st.button("🚪 Sair", use_container_width=True):
            st.session_state.authenticated = False
            st.rerun()

# Configuração da página
st.set_page_config(
    page_title="Dashboard Financeiro",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Verifica autenticação
check_authentication()

# Auto-refresh para atualizar automaticamente
st_autorefresh = st.empty()
with st_autorefresh:
    if st.button("🔄 Atualizar Dados", key="refresh_btn", help="Clique para atualizar ou aguarde 30s"):
        st.rerun()

# CSS COM MELHORIAS VISUAIS SUTIS E RESPONSIVO
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    * {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }
    
    /* MÉTRICAS DELICADAS E ELEGANTES */
    [data-testid="metric-container"] {
        background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
        padding: 0.8rem 1rem !important;
        border-radius: 16px;
        border: 1px solid rgba(30, 60, 114, 0.08);
        border-left: 3px solid #1e3c72;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05), 0 1px 2px rgba(0, 0, 0, 0.1);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        margin-bottom: 0.5rem;
    }
    
    [data-testid="metric-container"]:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.08), 0 2px 4px rgba(30, 60, 114, 0.1);
        border-left: 3px solid #2a5298;
    }
    
    /* Labels delicados */
    [data-testid="metric-container"] label {
        color: #6b7280 !important;
        font-size: 0.65rem !important;
        font-weight: 500 !important;
        text-transform: uppercase;
        letter-spacing: 0.8px;
        margin-bottom: 0.2rem !important;
        opacity: 0.9;
    }
    
    /* Valores das métricas menores - múltiplos seletores para garantir */
    [data-testid="metric-container"] > div > div:first-child,
    [data-testid="metric-container"] div[data-testid="metric-container"] > div:nth-child(2),
    [data-testid="metric-container"] > div:last-child > div:first-child,
    [data-testid="metric-container"] .metric-value {
        font-size: 0.65rem !important;
        font-weight: 600 !important;
        color: #1f2937 !important;
        line-height: 1.2 !important;
    }
    
    /* Delta menor */
    [data-testid="metric-container"] [data-testid="stMetricDelta"] {
        font-size: 0.7rem !important;
        font-weight: 500 !important;
    }
    
    /* Tabs melhoradas */
    .stTabs [data-baseweb="tab-list"] {
        background: white;
        border-radius: 10px;
        padding: 4px;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
        margin-bottom: 1rem;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border-radius: 8px;
        padding: 0.6rem 1rem !important;
        font-weight: 500;
        font-size: 0.85rem !important;
        color: #6b7280;
        border: none;
    }
    
    .stTabs [aria-selected="true"] {
        background: #1e3c72 !important;
        color: white !important;
    }
    
    /* Botões melhorados */
    .stButton > button {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.6rem 1.2rem;
        font-weight: 500;
        font-size: 0.85rem;
        transition: all 0.2s;
    }
    
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 8px rgba(30, 60, 114, 0.3);
    }
    
    /* Headers menores */
    h1 {
        font-size: 1.8rem !important;
        font-weight: 700 !important;
        color: #1f2937 !important;
    }
    
    h2 {
        font-size: 1.4rem !important;
        font-weight: 600 !important;
        color: #374151 !important;
    }
    
    h3 {
        font-size: 1.1rem !important;
        font-weight: 600 !important;
        color: #4b5563 !important;
    }
    
    /* Tabelas melhoradas */
    .dataframe {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
        font-size: 0.85rem !important;
        border: none !important;
    }
    
    .dataframe thead tr th {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%) !important;
        color: white !important;
        font-weight: 600 !important;
        font-size: 0.8rem !important;
        padding: 1rem 0.8rem !important;
        text-align: center !important;
        border: none !important;
    }
    
    .dataframe tbody tr {
        border-bottom: 1px solid #f1f5f9 !important;
    }
    
    .dataframe tbody tr:hover {
        background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%) !important;
        transform: translateY(-1px);
        transition: all 0.2s ease;
    }
    
    .dataframe tbody tr td {
        padding: 0.8rem !important;
        text-align: center !important;
        font-weight: 500 !important;
        border: none !important;
    }
    
    /* Estilo para valores monetários */
    .dataframe tbody tr td:contains("R$") {
        font-weight: 600 !important;
        color: #059669 !important;
    }
    
    /* Alternating row colors */
    .dataframe tbody tr:nth-child(even) {
        background: #f8fafc !important;
    }
    
    .dataframe tbody tr:nth-child(odd) {
        background: white !important;
    }
    
    /* Mobile responsivo e delicado */
    @media (max-width: 768px) {
        [data-testid="metric-container"] {
            padding: 0.6rem 0.8rem !important;
            margin-bottom: 0.4rem;
            border-radius: 12px;
        }
        
        [data-testid="metric-container"] label {
            font-size: 0.6rem !important;
            letter-spacing: 0.6px;
        }
        
        [data-testid="metric-container"] > div > div:first-child,
        [data-testid="metric-container"] div[data-testid="metric-container"] > div:nth-child(2),
        [data-testid="metric-container"] > div:last-child > div:first-child,
        [data-testid="metric-container"] .metric-value {
            font-size: 0.55rem !important;
            line-height: 1.1 !important;
        }
        
        [data-testid="metric-container"] [data-testid="stMetricDelta"] {
            font-size: 0.6rem !important;
        }
        
        h1 {
            font-size: 1.4rem !important;
        }
        
        h2 {
            font-size: 1.1rem !important;
        }
        
        .stTabs [data-baseweb="tab"] {
            padding: 0.4rem 0.7rem !important;
            font-size: 0.75rem !important;
        }
        
        .dataframe {
            font-size: 0.75rem !important;
        }
    }
    
    /* Classes originais mantidas */
    .categoria-principal {
        background: linear-gradient(135deg, #f0f2f6 0%, #e5e7eb 100%);
        border-left: 4px solid #1e3c72;
        padding: 0.8rem 1rem;
        margin: 0.5rem 0;
        font-weight: 600;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
    }
    
    .subcategoria {
        background: #f9fafb;
        border-left: 2px solid #d1d5db;
        padding: 0.5rem 1.5rem;
        margin: 0.25rem 0;
        font-size: 0.85em;
        border-radius: 4px;
    }
    
    .main-header {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        color: white;
        padding: 1.8rem;
        border-radius: 12px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 6px 20px rgba(30, 60, 114, 0.3);
    }
    
    .metric-card {
        background: white;
        padding: 1.3rem;
        border-radius: 12px;
        border-left: 4px solid #1e3c72;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin-bottom: 1rem;
        transition: all 0.2s;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 15px rgba(0, 0, 0, 0.15);
    }
    
    .category-card {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin: 0.5rem 0;
        box-shadow: 0 4px 8px rgba(30, 60, 114, 0.3);
        transition: all 0.2s;
    }
    
    .category-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 15px rgba(30, 60, 114, 0.4);
    }
    
    .success-badge {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        padding: 0.3rem 0.6rem;
        border-radius: 15px;
        font-size: 0.75rem;
        font-weight: 600;
        box-shadow: 0 2px 4px rgba(16, 185, 129, 0.3);
    }
    
    /* Background geral sutil */
    .main .block-container {
        background: linear-gradient(145deg, #f8fafc 0%, #f1f5f9 50%, #e2e8f0 100%);
        padding: 2rem 1rem;
    }
    
    /* Cards de informação melhorados */
    .stAlert {
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08) !important;
    }
    
    .stAlert[data-baseweb="notification"] {
        background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%) !important;
        color: #1e40af !important;
    }
    
    /* Gráficos com bordas arredondadas */
    .js-plotly-plot {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08);
        background: white;
    }
    
    /* Sidebar melhorada (se existir) */
    .css-1d391kg {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%) !important;
    }
    
    /* Info boxes melhoradas */
    .stInfo {
        background: linear-gradient(135deg, #e0f2fe 0%, #b3e5fc 100%) !important;
        border-left: 4px solid #0277bd !important;
        border-radius: 8px !important;
    }
    
    .stSuccess {
        background: linear-gradient(135deg, #e8f5e8 0%, #c8e6c9 100%) !important;
        border-left: 4px solid #388e3c !important;
        border-radius: 8px !important;
    }
    
    .stWarning {
        background: linear-gradient(135deg, #fff8e1 0%, #ffecb3 100%) !important;
        border-left: 4px solid #f57c00 !important;
        border-radius: 8px !important;
    }
    
    /* Refinamentos adicionais para elegância */
    .element-container {
        margin-bottom: 0.8rem !important;
    }
    
    /* Containers com mais respiro */
    .block-container {
        padding-top: 1.5rem !important;
        padding-bottom: 2rem !important;
    }
    
    /* Subtítulos mais elegantes */
    .stSubheader {
        color: #374151 !important;
        font-weight: 500 !important;
        margin-bottom: 1rem !important;
        font-size: 1.1rem !important;
    }
    
    /* Divisores sutis */
    hr {
        border: none !important;
        height: 1px !important;
        background: linear-gradient(90deg, transparent 0%, #e5e7eb 50%, transparent 100%) !important;
        margin: 1.5rem 0 !important;
    }
    
    /* Melhorias nos selectboxes */
    .stSelectbox > div > div > div {
        border-radius: 8px !important;
        border-color: #d1d5db !important;
    }
    
    /* Melhorias nos sliders se houver */
    .stSlider > div > div > div > div {
        background-color: #1e3c72 !important;
    }
    
    .warning-badge {
        background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);
        color: white;
        padding: 0.3rem 0.6rem;
        border-radius: 15px;
        font-size: 0.75rem;
        font-weight: 600;
        box-shadow: 0 2px 4px rgba(245, 158, 11, 0.3);
    }
    
    .danger-badge {
        background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
        color: white;
        padding: 0.3rem 0.6rem;
        border-radius: 15px;
        font-size: 0.75rem;
        font-weight: 600;
        box-shadow: 0 2px 4px rgba(239, 68, 68, 0.3);
    }
    
    .edit-section {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        padding: 1.5rem;
        border-radius: 12px;
        border: 2px dashed #adb5bd;
        margin: 1rem 0;
        transition: all 0.2s;
    }
    
    .edit-section:hover {
        border-color: #1e3c72;
        background: linear-gradient(135deg, #ffffff 0%, #f1f3f4 100%);
    }
    
    /* OCULTAR KEYS VISÍVEIS DE ELEMENTOS STREAMLIT */
    [data-testid*="key-"], 
    .streamlit-expanderHeader::before,
    .element-container[data-key*="key"]::before,
    .stExpander > div:first-child::before,
    [class*="key-"]::before,
    div[data-baseweb="accordion"] > div:first-child::before {
        display: none !important;
        visibility: hidden !important;
        opacity: 0 !important;
        height: 0 !important;
        width: 0 !important;
        position: absolute !important;
        left: -9999px !important;
    }
    
    /* Limpar qualquer conteúdo gerado automaticamente */
    .stExpander > div:first-child::before,
    .stExpander > div:first-child::after {
        content: "" !important;
        display: none !important;
    }
    
    /* Ocultar textos de depuração */
    div[class*="debug"], 
    span[class*="debug"], 
    div[id*="debug"],
    span[id*="debug"] {
        display: none !important;
    }
    
    /* CORRIGIR PROBLEMA DO keyboard_arrow_right */
    .streamlit-expanderHeader svg,
    .streamlit-expanderHeader::before,
    div[data-baseweb="accordion"] svg,
    div[data-baseweb="accordion"] .material-icons,
    div[data-baseweb="accordion"]::before {
        display: none !important;
        visibility: hidden !important;
    }
    
    /* Ocultar especificamente o texto keyboard_arrow_right */
    div:contains("keyboard_arrow_right"),
    span:contains("keyboard_arrow_right") {
        font-size: 0 !important;
        color: transparent !important;
        visibility: hidden !important;
    }
    
    /* Alternativa: ocultar qualquer texto que contenha 'keyboard' */
    *[title*="keyboard"],
    *[aria-label*="keyboard"],
    *:contains("keyboard_arrow") {
        display: none !important;
        font-size: 0 !important;
        color: transparent !important;
    }
    
    /* ESPECÍFICO PARA EXPANDERS - Ocultar setas completamente */
    div[data-baseweb="accordion"] button svg,
    div[data-baseweb="accordion"] button::before,
    div[data-baseweb="accordion"] button::after,
    .stExpander > div:first-child button svg,
    .stExpander > div:first-child button::before,
    .stExpander > div:first-child button::after {
        display: none !important;
        visibility: hidden !important;
        position: absolute !important;
        left: -9999px !important;
    }
    
    /* Remove qualquer overflow de texto dos expanders */
    div[data-baseweb="accordion"] button {
        overflow: hidden !important;
        text-overflow: clip !important;
        white-space: nowrap !important;
    }
    
</style>

<script>
// Função para remover elementos com keyboard_arrow_right
function removeKeyboardArrows() {
    // Aguarda um pouco para que o DOM seja renderizado
    setTimeout(function() {
        // Remove todos os elementos que contenham "keyboard_arrow_right"
        const elements = document.querySelectorAll('*');
        elements.forEach(function(element) {
            if (element.textContent && element.textContent.includes('keyboard_arrow_right')) {
                element.style.display = 'none';
                element.style.visibility = 'hidden';
                element.style.opacity = '0';
                element.remove();
            }
            // Também remove se for um atributo
            if (element.getAttribute('aria-label') && element.getAttribute('aria-label').includes('keyboard_arrow_right')) {
                element.style.display = 'none';
                element.remove();
            }
        });
        
        // Específico para SVGs e Material Icons
        const svgs = document.querySelectorAll('svg, .material-icons, [class*="icon"]');
        svgs.forEach(function(svg) {
            if (svg.innerHTML && svg.innerHTML.includes('keyboard_arrow_right')) {
                svg.style.display = 'none';
                svg.remove();
            }
        });
        
        // Remove elementos de expander que possam ter problemas
        const expanderButtons = document.querySelectorAll('div[data-baseweb="accordion"] button');
        expanderButtons.forEach(function(button) {
            const textNodes = Array.from(button.childNodes).filter(node => node.nodeType === Node.TEXT_NODE);
            textNodes.forEach(function(textNode) {
                if (textNode.textContent.includes('keyboard_arrow_right')) {
                    textNode.textContent = '';
                    textNode.remove();
                }
            });
        });
        
    }, 100);
}

// Executa a função quando a página carrega
document.addEventListener('DOMContentLoaded', removeKeyboardArrows);

// Executa novamente após mudanças no DOM
const observer = new MutationObserver(removeKeyboardArrows);
observer.observe(document.body, { childList: true, subtree: true });

// Executa também em intervalos para garantir
setInterval(removeKeyboardArrows, 1000);
</script>
""", unsafe_allow_html=True)

def formatar_moeda_br(valor):
    """Formata valor monetário no padrão brasileiro"""
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def get_status_badge(status):
    """Retorna badge colorido para status"""
    if status == 'Realizado':
        return f'<span class="success-badge">{status}</span>'
    elif status == 'Projetado':
        return f'<span class="warning-badge">{status}</span>'
    else:
        return f'<span class="danger-badge">{status}</span>'

@st.cache_data(ttl=30)
def carregar_dados():
    """Carrega todos os dados processados"""
    try:
        # Carrega despesas corrigidas
        df_despesas = pd.read_excel('despesas_processadas_correto.xlsx', sheet_name='Despesas_Completas')
        df_despesas['Data'] = pd.to_datetime(df_despesas['Data'])
        
        # Carrega receitas do sistema integrado
        df_receitas = pd.read_excel('sistema_financeiro_integrado.xlsx', sheet_name='Receitas')
        df_receitas['Data'] = pd.to_datetime(df_receitas['Data'], format='%d/%m/%Y')
        
        # Carrega fluxo de caixa
        df_fluxo = pd.read_excel('sistema_financeiro_integrado.xlsx', sheet_name='Fluxo_Diario')
        df_fluxo['Data'] = pd.to_datetime(df_fluxo['Data'], format='%d/%m/%Y')
        
        return df_despesas, df_receitas, df_fluxo
        
    except Exception as e:
        st.error(f"❌ Erro ao carregar dados: {e}")
        return None, None, None

def criar_overview_cards(df_despesas, df_receitas):
    """Cria cards de overview apenas com dados REALIZADOS"""
    
    col1, col2, col3, col4 = st.columns(4)
    
    # Receitas REALIZADAS apenas
    receitas_realizadas = df_receitas[df_receitas['Status'] == 'Realizado']['Valor'].sum()
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h3>Receitas Realizadas</h3>
            <h2 style="color: #28a745;">{formatar_moeda_br(receitas_realizadas)}</h2>
            <small>Apenas valores efetivados</small>
        </div>
        """, unsafe_allow_html=True)
    
    # Despesas REALIZADAS apenas
    despesas_realizadas = df_despesas[df_despesas['Status'] == 'Realizado']['Valor'].sum()
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <h3>Despesas Realizadas</h3>
            <h2 style="color: #dc3545;">{formatar_moeda_br(despesas_realizadas)}</h2>
            <small>Apenas valores efetivados</small>
        </div>
        """, unsafe_allow_html=True)
    
    # Resultado REALIZADO
    resultado_realizado = receitas_realizadas - despesas_realizadas
    cor_resultado = "#28a745" if resultado_realizado >= 0 else "#dc3545"
    
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <h3>Resultado Realizado</h3>
            <h2 style="color: {cor_resultado};">{formatar_moeda_br(resultado_realizado)}</h2>
            <small>Somente valores efetivados</small>
        </div>
        """, unsafe_allow_html=True)
    
    # Saldo Atual - dados reais do fluxo de caixa
    try:
        saldos_reais = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Saldos Adquirentes')
        saldo_atual = saldos_reais['Saldo Disponível'].sum()
    except:
        saldo_atual = 298373.39  # Fallback para o valor real mais recente
    
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <h3>🏦 Saldo Atual</h3>
            <h2 style="color: #2a5298;">{formatar_moeda_br(saldo_atual)}</h2>
            <small>Saldo atualizado</small>
        </div>
        """, unsafe_allow_html=True)

def mostrar_categorias_detalhadas(df_despesas):
    """Mostra análise detalhada das categorias"""
    
    st.header("🏷️ Análise Detalhada por Categoria")
    
    # Resumo por categoria
    resumo_cat = df_despesas.groupby('Categoria').agg({
        'Valor': 'sum',
        'Descricao': 'count',
        'Status': lambda x: (x == 'Realizado').sum()
    }).round(2)
    resumo_cat.columns = ['Total_Valor', 'Qtd_Itens', 'Qtd_Realizadas']
    resumo_cat['Qtd_Projetadas'] = resumo_cat['Qtd_Itens'] - resumo_cat['Qtd_Realizadas']
    resumo_cat = resumo_cat.sort_values('Total_Valor', ascending=False)
    
    # Cards por categoria
    for categoria, dados in resumo_cat.iterrows():
        with st.expander(f"{categoria} - {formatar_moeda_br(dados['Total_Valor'])}", expanded=False):
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total", formatar_moeda_br(dados['Total_Valor']))
            
            with col2:
                st.metric("Total Itens", int(dados['Qtd_Itens']))
            
            with col3:
                st.metric("✅ Realizadas", int(dados['Qtd_Realizadas']))
            
            # Detalhes dos itens desta categoria
            itens_categoria = df_despesas[df_despesas['Categoria'] == categoria].copy()
            itens_categoria['Status_Badge'] = itens_categoria['Status'].apply(get_status_badge)
            
            # Tabela formatada
            df_display = itens_categoria[['Descricao', 'Valor', 'Status', 'Mes', 'Data']].copy()
            df_display['Valor'] = df_display['Valor'].apply(formatar_moeda_br)
            df_display['Data'] = df_display['Data'].dt.strftime('%d/%m/%Y')
            
            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Descricao': st.column_config.TextColumn('Descrição', width='large'),
                    'Valor': st.column_config.TextColumn('Valor'),
                    'Status': st.column_config.TextColumn('Status'),
                    'Mes': st.column_config.TextColumn('Mês'),
                    'Data': st.column_config.TextColumn('Data')
                }
            )

def mostrar_graficos_avancados(df_despesas, df_receitas):
    """Mostra gráficos avançados e interativos"""
    
    st.header("Análises Visuais Avançadas")
    
    col1, col2 = st.columns(2)
    
    # Gráfico 1: Receitas vs Despesas por mês
    with col1:
        st.subheader("Receitas vs Despesas Mensais")
        
        # Prepara dados mensais
        df_receitas['Mes'] = df_receitas['Data'].dt.strftime('%b/%Y')
        df_despesas['Mes'] = df_despesas['Data'].dt.strftime('%b/%Y')
        
        receitas_mes = df_receitas.groupby(['Mes', 'Status'])['Valor'].sum().reset_index()
        despesas_mes = df_despesas.groupby('Mes')['Valor'].sum().reset_index()
        
        fig = make_subplots(
            rows=2, cols=1,
            subplot_titles=('Receitas', 'Despesas'),
            vertical_spacing=0.1
        )
        
        # Receitas
        for status in receitas_mes['Status'].unique():
            data = receitas_mes[receitas_mes['Status'] == status]
            cor = '#1f4e79' if status == 'Realizado' else '#87ceeb'
            fig.add_trace(
                go.Bar(name=f'Receitas {status}', x=data['Mes'], y=data['Valor'],
                       marker_color=cor, showlegend=True),
                row=1, col=1
            )
        
        # Despesas
        fig.add_trace(
            go.Bar(name='Despesas', x=despesas_mes['Mes'], y=despesas_mes['Valor'],
                   marker_color='#dc3545', showlegend=True),
            row=2, col=1
        )
        
        fig.update_layout(height=600, title_text="Comparativo Mensal")
        st.plotly_chart(fig, use_container_width=True)
    
    # Gráfico 2: Pizza das categorias
    with col2:
        st.subheader("🥧 Distribuição por Categoria")
        
        cat_dados = df_despesas.groupby('Categoria')['Valor'].sum().reset_index()
        
        fig = px.pie(
            cat_dados,
            values='Valor',
            names='Categoria',
            title="Despesas por Categoria",
            color_discrete_sequence=px.colors.qualitative.Set3
        )
        
        fig.update_traces(
            textposition='inside',
            textinfo='percent+label',
            hovertemplate='<b>%{label}</b><br>Valor: R$ %{value:,.2f}<br>Percentual: %{percent}<extra></extra>'
        )
        
        fig.update_layout(height=600)
        st.plotly_chart(fig, use_container_width=True)

def mostrar_fluxo_caixa_integrado():
    """Fluxo de caixa integrado com dados da matriz de projeções"""
    
    st.header("Fluxo de Caixa Integrado")
    
    try:
        # Carrega dados da matriz de projeções
        df_receitas_mensais = pd.read_excel('projecoes_5_anos.xlsx', sheet_name='Receitas_Mensais')
        df_despesas_mensais = pd.read_excel('projecoes_5_anos.xlsx', sheet_name='Despesas_Detalhadas')
        df_custos_mensais = pd.read_excel('projecoes_5_anos.xlsx', sheet_name='Custos_Detalhados')
        
        # Processa dados mensais
        fluxo_mensal = []
        
        # Carrega dados reais primeiro para integrar ao fluxo
        saldo_inicial_real = 0
        a_receber_asaas = 0
        
        try:
            # Usa a planilha correta do relatório de fluxo de caixa
            resumo_real = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Resumo')
            
            # Busca saldo do Asaas
            saldo_asaas = resumo_real[resumo_real['Métrica'] == 'Saldo Asaas']['Valor'].iloc[0]
            # Busca saldo do Pagar.me
            saldo_pagarme = resumo_real[resumo_real['Métrica'] == 'Saldo Pagar.me']['Valor'].iloc[0]
            # Soma os saldos
            saldo_inicial_real = saldo_asaas + saldo_pagarme
            
            # Busca valor a receber do Asaas
            a_receber_asaas = resumo_real[resumo_real['Métrica'] == 'A Receber Asaas']['Valor'].iloc[0]
            
        except Exception as e:
            saldo_inicial_real = 298373.39  # Saldo real atual (265125.61 + 33247.78)
            a_receber_asaas = 243462.42     # Valor a receber real
            st.warning("⚠️ Usando dados reais salvos - arquivo de fluxo pode estar indisponível")
        
        # Agrupa receitas por mês
        receitas_por_mes = df_receitas_mensais.groupby('Mes_Ano')['Valor'].sum()
        
        # Adiciona cronograma de recebimentos futuro às receitas mensais
        if a_receber_asaas > 0:
            valor_mensal_recebimento = a_receber_asaas / 3  # Distribui em 3 meses
            
            # Adiciona recebimentos aos próximos meses
            for i in range(3):
                mes_recebimento = pd.Timestamp.now() + pd.DateOffset(months=i+1)
                mes_ano_recebimento = mes_recebimento.strftime('%Y-%m')
                
                # Adiciona às receitas mensais
                if mes_ano_recebimento in receitas_por_mes.index:
                    receitas_por_mes[mes_ano_recebimento] += valor_mensal_recebimento
                else:
                    receitas_por_mes[mes_ano_recebimento] = valor_mensal_recebimento
                    
            st.success(f"Saldo inicial: {formatar_moeda_br(saldo_inicial_real)} | A receber: {formatar_moeda_br(a_receber_asaas)} (distribuído em 3 meses)")
        else:
            st.info(f"Saldo inicial atualizado: {formatar_moeda_br(saldo_inicial_real)}")
        
        # Agrupa despesas por mês
        despesas_por_mes = df_despesas_mensais.groupby('Mes_Ano')['Valor'].sum()
        
        # Agrupa custos por mês
        custos_por_mes = df_custos_mensais.groupby('Mes_Ano')['Valor'].sum()
        
        # Define saldo acumulado inicial
        saldo_acumulado = saldo_inicial_real
        
        # Cria dados do fluxo
        todos_meses = sorted(set(list(receitas_por_mes.index) + list(despesas_por_mes.index) + list(custos_por_mes.index)))
        
        for mes in todos_meses:
            receitas = receitas_por_mes.get(mes, 0)
            despesas = despesas_por_mes.get(mes, 0)  
            custos = custos_por_mes.get(mes, 0)
            
            fluxo_liquido = receitas - custos - despesas
            saldo_acumulado += fluxo_liquido
            
            fluxo_mensal.append({
                'Mes_Ano': mes,
                'Data': pd.to_datetime(f"{mes}-01"),
                'Receitas': receitas,
                'Custos': custos,
                'Despesas': despesas,
                'Fluxo_Liquido': fluxo_liquido,
                'Saldo_Acumulado': saldo_acumulado
            })
        
        df_fluxo = pd.DataFrame(fluxo_mensal)
        
        # Métricas principais
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Saldo Inicial", formatar_moeda_br(75015.85))
        
        with col2:
            saldo_final = df_fluxo['Saldo_Acumulado'].iloc[-1] if not df_fluxo.empty else 75015.85
            st.metric("Saldo Final", formatar_moeda_br(saldo_final))
        
        with col3:
            menor_saldo = df_fluxo['Saldo_Acumulado'].min() if not df_fluxo.empty else 75015.85
            st.metric("⚠️ Menor Saldo", formatar_moeda_br(menor_saldo))
        
        with col4:
            variacao = ((saldo_final - 75015.85) / 75015.85) * 100 if not df_fluxo.empty else 0
            st.metric("Variação", f"{variacao:.1f}%")
        
        # Gráfico de evolução
        st.subheader("Evolução do Fluxo de Caixa")
        
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=df_fluxo['Data'],
            y=df_fluxo['Saldo_Acumulado'],
            mode='lines+markers',
            name='Saldo Acumulado',
            line=dict(color='#2a5298', width=3),
            marker=dict(size=6)
        ))
        
        # Linha de saldo inicial
        fig.add_hline(y=75015.85, line_dash="dot", line_color="gray", 
                      annotation_text="Saldo Inicial")
        
        # Linha de alerta
        fig.add_hline(y=20000, line_dash="dash", line_color="red",
                      annotation_text="Nível de Alerta")
        
        fig.update_layout(
            title="Projeção de Saldo Acumulado",
            xaxis_title="Período",
            yaxis_title="Saldo (R$)",
            yaxis_tickformat=',.0f',
            height=500
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Tabela resumida
        st.subheader("Resumo Mensal")
        
        df_display = df_fluxo.copy()
        df_display['Mes_Ano'] = df_display['Mes_Ano']
        for col in ['Receitas', 'Custos', 'Despesas', 'Fluxo_Liquido', 'Saldo_Acumulado']:
            df_display[col] = df_display[col].apply(formatar_moeda_br)
        
        st.dataframe(
            df_display[['Mes_Ano', 'Receitas', 'Custos', 'Despesas', 'Fluxo_Liquido', 'Saldo_Acumulado']],
            use_container_width=True,
            hide_index=True
        )
        
    except Exception as e:
        st.error(f"❌ Erro ao carregar fluxo de caixa: {e}")
        st.info("💡 Execute primeiro a integração da matriz de projeções")

def mostrar_fluxo_caixa(df_fluxo):
    """Mostra análise do fluxo de caixa"""
    
    st.header("Análise de Fluxo de Caixa")
    
    # Métricas principais
    col1, col2, col3, col4 = st.columns(4)
    
    # Saldo inicial atualizado com dados reais
    try:
        saldos_reais = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Saldos Adquirentes')
        saldo_disponivel = saldos_reais['Saldo Disponível'].sum()
        saldo_a_receber = saldos_reais['A Receber'].sum()
        saldo_inicial = saldo_disponivel + saldo_a_receber
    except:
        saldo_inicial = 653141.31
    
    saldo_final = df_fluxo['Saldo_Acumulado'].iloc[-1] if not df_fluxo.empty else saldo_inicial
    menor_saldo = df_fluxo['Saldo_Acumulado'].min() if not df_fluxo.empty else saldo_inicial
    data_menor = df_fluxo[df_fluxo['Saldo_Acumulado'] == menor_saldo]['Data'].iloc[0] if not df_fluxo.empty else datetime.now()
    
    with col1:
        st.metric("Saldo Inicial", formatar_valor_compacto(saldo_inicial), help="Saldo em 04/09/2025")
    
    with col2:
        st.metric("Saldo Final Projetado", formatar_valor_compacto(saldo_final))
    
    with col3:
        st.metric("Menor Saldo", formatar_valor_compacto(menor_saldo), 
                 f"{data_menor.strftime('%d/%m/%Y')}")
    
    with col4:
        variacao = ((saldo_final - saldo_inicial) / saldo_inicial) * 100
        st.metric("Variação", f"{variacao:.1f}%")
    
    # Gráfico de fluxo de caixa
    st.subheader("Evolução do Saldo")
    
    fig = go.Figure()
    
    # Linha do saldo
    fig.add_trace(go.Scatter(
        x=df_fluxo['Data'],
        y=df_fluxo['Saldo_Acumulado'],
        mode='lines+markers',
        name='Saldo',
        line=dict(color='#2a5298', width=2),
        marker=dict(size=4)
    ))
    
    # Linha de referência zero
    fig.add_hline(y=0, line_dash="dash", line_color="red", opacity=0.5)
    
    # Linha de saldo mínimo seguro (R$ 20.000)
    fig.add_hline(y=20000, line_dash="dot", line_color="orange", opacity=0.5,
                  annotation_text="Saldo Mínimo Seguro")
    
    fig.update_layout(
        title="Projeção de Fluxo de Caixa",
        xaxis_title="Data",
        yaxis_title="Saldo (R$)",
        yaxis_tickformat=',.0f',
        height=500,
        hovermode='x unified'
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Tabela resumida mensal
    st.subheader("Resumo Mensal")
    
    # Agrupa por mês
    df_mensal = df_fluxo.copy()
    df_mensal['Mes'] = df_mensal['Data'].dt.strftime('%b/%Y')
    
    resumo_mensal = df_mensal.groupby('Mes').agg({
        'Receitas': 'sum',
        'Despesas': 'sum',
        'Fluxo_Liquido': 'sum',
        'Saldo_Acumulado': 'last'
    }).round(2)
    
    # Formata valores
    for col in resumo_mensal.columns:
        resumo_mensal[col] = resumo_mensal[col].apply(formatar_moeda_br)
    
    st.dataframe(resumo_mensal, use_container_width=True)

def carregar_projecoes_plurianuais():
    """Carrega projeções de 5 anos usando dados REAIS da Matriz financeira.xlsx"""
    
    try:
        # Carrega dados reais da matriz financeira
        df_matriz = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
        
        # Identifica colunas de data
        colunas_data = []
        for col in df_matriz.columns:
            if isinstance(col, (pd.Timestamp, datetime)):
                colunas_data.append(col)
        
        # Prepara dados consolidados por ano (usando a mesma lógica correta)
        dados_anuais = {}
        for ano in range(2025, 2031):
            dados_anuais[ano] = {
                'receitas_total': 0,
                'custos_total': 0,
                'despesas_total': 0,
                'impostos_total': 0
            }
        
        # Usa EXATAMENTE a mesma lógica da tabela unificada
        # Processa dados mensais e soma por ano
        for col in colunas_data:
            if isinstance(col, (datetime, pd.Timestamp)):
                ano = col.year
                if ano in dados_anuais:
                    
                    # RECEITAS - usa TOTAL RECEITAS direto
                    receitas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL RECEITAS']
                    if not receitas_linha.empty:
                        valor_receita = converter_para_float_seguro(receitas_linha[col].iloc[0])
                        dados_anuais[ano]['receitas_total'] += valor_receita
                    
                    # CUSTOS - usa TOTAL CUSTOS direto
                    custos_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL CUSTOS']
                    if not custos_linha.empty:
                        valor_custo = converter_para_float_seguro(custos_linha[col].iloc[0])
                        dados_anuais[ano]['custos_total'] += valor_custo
                    
                    # DESPESAS - usa TOTAL DESPESAS direto
                    despesas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL DESPESAS']
                    if not despesas_linha.empty:
                        valor_despesa = converter_para_float_seguro(despesas_linha[col].iloc[0])
                        dados_anuais[ano]['despesas_total'] += valor_despesa
                    
                    # IMPOSTOS - calcula baseado nas receitas mensais (Lucro Presumido)
                    if col >= pd.to_datetime('2025-10-01'):
                        if 'valor_receita' in locals() and valor_receita > 0:
                            receitas_servicos = valor_receita * 0.4
                            receitas_livros = valor_receita * 0.6
                            
                            # USA A FUNÇÃO CENTRALIZADA PARA GARANTIR CONSISTÊNCIA
                            impostos_mes_dict = calcular_impostos_brasileiros(receitas_servicos, receitas_livros, col)
                            impostos_mes_total = sum(impostos_mes_dict.values())
                            dados_anuais[ano]['impostos_total'] += impostos_mes_total
        
        # Converte para DataFrame no formato esperado
        projecoes = []
        for ano in range(2025, 2031):
            if dados_anuais[ano]['receitas_total'] > 0:  # Só inclui anos com dados
                resultado_liquido = (dados_anuais[ano]['receitas_total'] - 
                                    dados_anuais[ano]['custos_total'] - 
                                    dados_anuais[ano]['despesas_total'] - 
                                    dados_anuais[ano]['impostos_total'])
                
                projecoes.append({
                    'Ano': ano,
                    'Receitas_Brutas': dados_anuais[ano]['receitas_total'],
                    'Custos_Operacionais': dados_anuais[ano]['custos_total'],
                    'Despesas_Totais': dados_anuais[ano]['despesas_total'],
                    'Impostos_Totais': dados_anuais[ano]['impostos_total'],
                    'Resultado_Liquido': resultado_liquido
                })
        
        return pd.DataFrame(projecoes)
        
    except Exception as e:
        st.error(f"Erro ao carregar projeções reais: {e}")
        # Fallback para dados básicos se houver erro
        anos = list(range(2025, 2031))
        dados_exemplo = []
        
        for ano in anos:
            dados_exemplo.append({
                'Ano': ano,
                'Receitas_Brutas': 1000000 * (1.15 ** (ano - 2025)),
                'Custos_Operacionais': 50000 * (1.15 ** (ano - 2025)),
                'Despesas_Totais': 400000 * (1.15 ** (ano - 2025)),
                'Impostos_Totais': 150000 * (1.15 ** (ano - 2025)),
                'Resultado_Liquido': 400000 * (1.15 ** (ano - 2025))
            })
        
        return pd.DataFrame(dados_exemplo)

def mostrar_projecoes_5_anos():
    """Seção de projeções plurianuais"""
    
    st.header("Projeções")
    
    st.info("💡 **Aguardando dados reais das projeções.** Esta seção será atualizada com suas projeções de 5 anos.")
    
    # Carrega dados de projeção
    df_proj = carregar_projecoes_plurianuais()
    
    if df_proj.empty:
        st.warning("Nenhum dado de projeção encontrado. Adicione o arquivo 'projecoes_5_anos.xlsx'")
        return
    
    # Seletor de visualização
    st.subheader("Modo de Visualização")
    
    col1, col2 = st.columns(2)
    
    with col1:
        modo_visualizacao = st.radio(
            "Escolha o modo:",
            ["Detalhado", "Consolidado", "Matriz", "Tabela Unificada"],
            horizontal=False,
            key="modo_visualizacao_projecoes"
        )
    
    with col2:
        anos_disponiveis = sorted(df_proj['Ano'].unique())
        if modo_visualizacao == "Detalhado":
            ano_selecionado = st.selectbox("Selecione o ano:", anos_disponiveis, key="ano_selecionado_projecoes")
        
        # Botão para forçar atualização dos dados
        if st.button("🔄 Atualizar Matriz", help="Recarrega dados da Matriz financeira.xlsx", key="btn_atualizar_matriz"):
            st.cache_data.clear()
            st.rerun()
    
    # Métricas consolidadas (5 anos)
    st.subheader("Resumo Consolidado (5 Anos)")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        receitas_total = df_proj['Receitas_Brutas'].sum()
        st.metric("Receitas Totais", formatar_valor_compacto(receitas_total))
    
    with col2:
        custos_total = df_proj['Custos_Operacionais'].sum()
        st.metric("Custos Totais", formatar_valor_compacto(custos_total))
    
    with col3:
        despesas_total = df_proj['Despesas_Totais'].sum()
        st.metric("Despesas Totais", formatar_valor_compacto(despesas_total))
    
    with col4:
        resultado_total = df_proj['Resultado_Liquido'].sum()
        cor_resultado = "#28a745" if resultado_total >= 0 else "#dc3545"
        st.metric("Resultado Líquido", formatar_valor_compacto(resultado_total))
    
    # Visualização por modo
    if modo_visualizacao == "Detalhado":
        mostrar_detalhes_ano(df_proj, ano_selecionado)
    elif modo_visualizacao == "Matriz":
        mostrar_matriz_detalhada()
    elif modo_visualizacao == "Tabela Unificada":
        mostrar_tabela_unificada(df_proj)
    elif modo_visualizacao == "Consolidado":
        mostrar_consolidado_5_anos(df_proj)

def mostrar_detalhes_ano(df_proj, ano):
    """Mostra detalhes de um ano específico"""
    
    st.subheader(f"Detalhes do Ano {ano}")
    
    # Filtra dados do ano
    dados_ano = df_proj[df_proj['Ano'] == ano]
    
    if dados_ano.empty:
        st.error(f"❌ Não foram encontrados dados para o ano {ano}")
        return
    
    row = dados_ano.iloc[0]
    
    # Cards de métricas do ano
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h4>Receitas {ano}</h4>
            <h3 style="color: #28a745;">{formatar_moeda_br(row['Receitas_Brutas'])}</h3>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <h4>Custos {ano}</h4>
            <h3 style="color: #fd7e14;">{formatar_moeda_br(row['Custos_Operacionais'])}</h3>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <h4>Despesas {ano}</h4>
            <h3 style="color: #dc3545;">{formatar_moeda_br(row['Despesas_Totais'])}</h3>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        cor = "#28a745" if row['Resultado_Liquido'] >= 0 else "#dc3545"
        st.markdown(f"""
        <div class="metric-card">
            <h4>Resultado {ano}</h4>
            <h3 style="color: {cor};">{formatar_moeda_br(row['Resultado_Liquido'])}</h3>
        </div>
        """, unsafe_allow_html=True)
    
    # Gráfico de composição do resultado
    st.subheader(f"🥧 Composição Financeira - {ano}")
    
    valores = [row['Receitas_Brutas'], row['Custos_Operacionais'], row['Despesas_Totais']]
    labels = ['Receitas', 'Custos', 'Despesas']
    cores = ['#28a745', '#fd7e14', '#dc3545']
    
    fig = go.Figure(data=[go.Bar(
        x=labels,
        y=valores,
        marker_color=cores,
        text=[formatar_moeda_br(v) for v in valores],
        textposition='auto'
    )])
    
    fig.update_layout(
        title=f"Composição Financeira - {ano}",
        yaxis_title="Valores (R$)",
        showlegend=False,
        height=400
    )
    
    st.plotly_chart(fig, use_container_width=True)

def converter_para_float_seguro(x):
    """Converte valor para float tratando formatos brasileiros"""
    try:
        if pd.isna(x) or x == '' or x is None:
            return 0.0
        
        # Se já é número, retorna
        if isinstance(x, (int, float)):
            return float(x)
        
        # Converte string para float tratando formatos brasileiros
        valor_str = str(x).strip()
        
        # Remove espaços, tabs, quebras de linha
        valor_str = ''.join(valor_str.split())
        
        # Trata casos especiais
        if valor_str == '' or valor_str == '-' or valor_str == ';':
            return 0.0
        
        # Remove caracteres não numéricos exceto vírgula e ponto
        valor_limpo = ''
        for char in valor_str:
            if char.isdigit() or char in ',.':
                valor_limpo += char
        
        if not valor_limpo:
            return 0.0
        
        # Trata formato brasileiro (vírgula como decimal)
        if ',' in valor_limpo and '.' in valor_limpo:
            # Ex: 1.234,56
            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
        elif ',' in valor_limpo:
            # Ex: 1234,56
            valor_limpo = valor_limpo.replace(',', '.')
        
        return float(valor_limpo)
    
    except (ValueError, TypeError):
        return 0.0

def formatar_valor_limpo(x):
    """Formata valor de forma limpa, apenas números"""
    try:
        valor = converter_para_float_seguro(x)
        if valor == 0:
            return '-'
        # Formato simplificado: 123.456 (sem R$)
        if valor >= 1000000:  # Milhões
            return f"{valor/1000000:.1f}M"
        elif valor >= 1000:   # Milhares
            return f"{valor/1000:.0f}k"
        else:
            return f"{valor:,.0f}".replace(',', '.')
    except (ValueError, TypeError):
        return '-'

def formatar_resultado_limpo(x):
    """Formata resultado líquido de forma limpa com indicador"""
    try:
        valor = converter_para_float_seguro(x)
        if valor == 0:
            return '-'
        
        # Formato simplificado com indicador
        if valor >= 1000000:  # Milhões
            valor_str = f"{valor/1000000:.1f}M"
        elif valor >= 1000:   # Milhares
            valor_str = f"{valor/1000:.0f}k"
        else:
            valor_str = f"{valor:,.0f}".replace(',', '.')
            
        return f"{'✓' if valor >= 0 else '✗'} {valor_str}"
    except (ValueError, TypeError):
        return '-'

def calcular_impostos_brasileiros(receitas_servicos, receitas_livros, mes_data):
    """Calcula impostos brasileiros - Lucro Presumido (40% serviços digitais, 60% venda de livros)"""
    impostos = {}
    
    # Aplicar impostos a partir de outubro/2025
    if mes_data < datetime(2025, 10, 1):
        return impostos
    
    # Receita total
    receita_total = receitas_servicos + receitas_livros
    
    # 1. ISS (sobre serviços digitais apenas) - 5%
    iss = receitas_servicos * 0.02
    impostos['ISS'] = iss
    
    # 2. PIS (sobre faturamento total) - 0.65% no lucro presumido
    pis = receita_total * 0.0065
    impostos['PIS'] = pis
    
    # 3. COFINS (sobre faturamento total) - 3% no lucro presumido
    cofins = receita_total * 0.03
    impostos['COFINS'] = cofins
    
    # 4. Base de cálculo IRPJ/CSLL no Lucro Presumido:
    # - Serviços digitais: presunção de 32% sobre a receita
    # - Venda de livros: presunção de 1,6% sobre a receita (reduzida conforme lei)
    base_servicos = receitas_servicos * 0.32
    base_livros = receitas_livros * 0.016  # 1,6% para livros
    base_irpj_csll = base_servicos + base_livros
    
    # 5. IRPJ - 15% sobre a base + adicional de 10% sobre o que exceder R$ 20.000/mês
    if base_irpj_csll > 20000:
        irpj = (20000 * 0.15) + ((base_irpj_csll - 20000) * 0.10)  # Adicional é 10%, não 25%
    else:
        irpj = base_irpj_csll * 0.15
    impostos['IRPJ'] = irpj
    
    # 6. CSLL - 9% sobre a base de cálculo
    csll = base_irpj_csll * 0.09
    impostos['CSLL'] = csll
    
    return impostos

def processar_dados_organizados():
    """Organiza dados da matriz por seções com subcategorias e totais + impostos brasileiros"""
    df_matriz = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
    colunas_data = [col for col in df_matriz.columns if isinstance(col, datetime)]
    
    # Organizar dados por categorias principais
    secoes = {
        'RECEITAS': [],
        'CUSTOS': [],
        'IMPOSTOS_FEDERAIS': [],  # CSLL e IRPJ
        'DESPESAS': [],
        'RESULTADO': []
    }
    
    categoria_atual = None
    for _, row in df_matriz.iterrows():
        cat = row['Categoria']
        
        if pd.isna(cat) or str(cat).strip() == '':
            continue
            
        cat_str = str(cat).strip()
        
        # Identifica headers de seções
        if cat_str in ['RECEITAS', 'CUSTOS', 'DESPESAS']:
            categoria_atual = cat_str
        elif cat_str == 'RESULTADO LÍQUIDO':
            categoria_atual = 'RESULTADO'
        elif cat_str in ['TOTAL RECEITAS', 'TOTAL CUSTOS', 'TOTAL DESPESAS']:
            continue  # Vai ser calculado automaticamente
        # Filtra impostos das despesas
        elif categoria_atual == 'DESPESAS' and ('imposto' in cat_str.lower() or 'das' in cat_str.lower()):
            continue  # Remove impostos das despesas, serão calculados automaticamente
        # Adiciona itens às seções apropriadas
        elif categoria_atual and categoria_atual in secoes:
            secoes[categoria_atual].append(row)
    
    # Calcular impostos brasileiros - UMA LINHA POR IMPOSTO
    impostos_por_tributo = {'ISS': {}, 'PIS': {}, 'COFINS': {}, 'IRPJ': {}, 'CSLL': {}}
    
    # Inicializa todas as colunas com 0
    for tributo in impostos_por_tributo:
        for col in colunas_data:
            impostos_por_tributo[tributo][col] = 0.0
    
    # Calcula impostos para cada mês
    for col in colunas_data:
        # Estimar receitas: 40% serviços digitais, 60% venda de livros
        receitas_totais = 0
        for row in secoes['RECEITAS']:
            valor_receita = converter_para_float_seguro(row[col])
            receitas_totais += valor_receita
        
        # Divisão conforme modelo de negócio
        receitas_servicos = receitas_totais * 0.4  # 40% serviços digitais
        receitas_livros = receitas_totais * 0.6    # 60% venda de livros
        
        impostos_mes = calcular_impostos_brasileiros(receitas_servicos, receitas_livros, col)
        
        # Atribui valores calculados ao mês específico
        for tributo, valor in impostos_mes.items():
            if valor > 0:
                impostos_por_tributo[tributo][col] = valor
    
    # Cria linhas únicas para cada imposto
    impostos_custos = []
    impostos_federais = []
    
    # ISS, PIS, COFINS para CUSTOS
    for tributo in ['ISS', 'PIS', 'COFINS']:
        if any(impostos_por_tributo[tributo].values()):  # Se tem algum valor > 0
            linha_imposto = pd.Series({
                'Categoria': f'{tributo}',
                'Tipo': 'Imposto',
                **impostos_por_tributo[tributo]
            }, name=tributo)
            impostos_custos.append(linha_imposto)
    
    # IRPJ, CSLL para IMPOSTOS_FEDERAIS
    for tributo in ['IRPJ', 'CSLL']:
        if any(impostos_por_tributo[tributo].values()):  # Se tem algum valor > 0
            linha_imposto = pd.Series({
                'Categoria': f'{tributo}',
                'Tipo': 'Imposto Federal',
                **impostos_por_tributo[tributo]
            }, name=tributo)
            impostos_federais.append(linha_imposto)
    
    # Adiciona impostos às seções
    secoes['CUSTOS'].extend(impostos_custos)
    secoes['IMPOSTOS_FEDERAIS'] = impostos_federais
    
    return secoes, colunas_data

def organizar_despesas_com_hierarquia(dados, colunas_data):
    """Organiza despesas por categoria com hierarquia visual"""
    if not dados:
        return []
    
    # Processa os dados sequencialmente
    linhas_organizadas = []
    categoria_atual = None
    categoria_dados = {}
    subcategorias = []
    
    for row in dados:
        categoria = str(row['Categoria']).strip()
        
        # Identifica se é subcategoria (começa com "  -> " ou "  →")
        if categoria.startswith('  -> ') or categoria.startswith('  →'):
            # É subcategoria
            subcategoria = categoria.replace('  -> ', '').replace('  → ', '').strip()
            
            # Adiciona subcategoria
            linha_item = {
                'Categoria': f'    {subcategoria}',
                'Tipo': 'Item',
                'nivel': 'item'
            }
            # Adiciona valores das colunas
            for col in colunas_data:
                linha_item[col] = row[col]
            subcategorias.append(linha_item)
            
            # Soma para a categoria principal APENAS se a categoria pai não tem valores próprios
            if categoria_atual:
                for col in colunas_data:
                    if col not in categoria_dados:
                        categoria_dados[col] = 0
                    # Só soma subcategoria se categoria pai tem valor zero
                    if categoria_dados[col] == 0:
                        valor = converter_para_float_seguro(row[col])
                        categoria_dados[col] += valor
        else:
            # Se há categoria anterior, finaliza ela
            if categoria_atual:
                # Adiciona categoria principal
                linha_categoria = {
                    'Categoria': categoria_atual,
                    'Tipo': 'CATEGORIA',
                    'nivel': 'categoria',
                    **categoria_dados
                }
                linhas_organizadas.append(linha_categoria)
                
                # Adiciona subcategorias
                linhas_organizadas.extend(subcategorias)
            
            # Inicia nova categoria
            categoria_atual = categoria
            categoria_dados = {}
            subcategorias = []
            
            # Adiciona valores da categoria principal
            for col in colunas_data:
                valor = converter_para_float_seguro(row[col])
                categoria_dados[col] = valor
                # Se categoria pai tem valor, não soma subcategorias depois
    
    # Finaliza última categoria
    if categoria_atual:
        linha_categoria = {
            'Categoria': categoria_atual,
            'Tipo': 'CATEGORIA',
            'nivel': 'categoria',
            **categoria_dados
        }
        linhas_organizadas.append(linha_categoria)
        linhas_organizadas.extend(subcategorias)
    
    return linhas_organizadas

def criar_tabela_despesas_hierarquica(dados, titulo, colunas_data, cor_emoji=""):
    """Cria tabela de despesas com hierarquia visual"""
    if not dados:
        return
        
    st.markdown(f"### {cor_emoji} **{titulo}**")
    
    # Organiza com hierarquia
    linhas_organizadas = organizar_despesas_com_hierarquia(dados, colunas_data)
    
    if linhas_organizadas:
        df_display = pd.DataFrame(linhas_organizadas)
        
        # Formata valores e aplica estilo para categorias principais
        for col in colunas_data:
            col_nome = col.strftime('%m/%y')
            df_display[col_nome] = df_display.apply(
                lambda row: formatar_valor_limpo(row[col]) if row['nivel'] == 'item' 
                else formatar_valor_limpo(row[col]) if row['nivel'] == 'categoria' and converter_para_float_seguro(row[col]) > 0
                else '-', axis=1
            )
        
        # Remove colunas desnecessárias para exibição
        colunas_exibir = ['Categoria'] + [col.strftime('%m/%y') for col in colunas_data]
        df_display = df_display[colunas_exibir]
        
        # Define as categorias principais (agora 8 com Produto e Equipe)
        categorias_principais = ['Marketing e Comercial', 'Produto', 'Impostos', 'Juridico', 'Pro labore', 'Serviços', 'Equipe', 'Sistemas']
        
        # Aplica formatação específica para as 5 categorias
        df_display_formatado = df_display.copy()
        
        def formatar_categoria(categoria):
            # Remove espaços iniciais para comparação
            categoria_limpa = categoria.strip()
            if categoria_limpa in categorias_principais:
                return f"🔸 {categoria_limpa}"
            else:
                return categoria
        
        df_display_formatado['Categoria'] = df_display_formatado['Categoria'].apply(formatar_categoria)
        
        # Exibe tabela limpa com ícones
        st.dataframe(
            df_display_formatado,
            use_container_width=True,
            height=min(len(df_display_formatado) * 35 + 50, 500),
            column_config={
                "Categoria": st.column_config.TextColumn("Categoria", width="large")
            }
        )
        
        # Usa o TOTAL DESPESAS direto da planilha original
        try:
            df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
            linha_total_despesas = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
            
            total_secao = {}
            if not linha_total_despesas.empty:
                # Usa valores diretos da planilha
                for col in colunas_data:
                    valor_total = linha_total_despesas[col].iloc[0]
                    col_nome = col.strftime('%m/%y')
                    if pd.notna(valor_total) and valor_total != 0:
                        total_secao[col_nome] = formatar_valor_limpo(float(valor_total))
                    else:
                        total_secao[col_nome] = '-'
            else:
                # Fallback: calcula manualmente se não encontrar
                for col in colunas_data:
                    total_col = 0
                    for row in dados:
                        categoria = str(row['Categoria']).strip()
                        # Soma apenas se NÃO é subcategoria
                        if not categoria.startswith('  -> ') and not categoria.startswith('  →'):
                            valor = converter_para_float_seguro(row[col])
                            total_col += valor
                    col_nome = col.strftime('%m/%y')
                    total_secao[col_nome] = formatar_valor_limpo(total_col) if total_col != 0 else '-'
        except:
            # Se der erro, calcula manualmente
            total_secao = {}
            for col in colunas_data:
                total_col = 0
                for row in dados:
                    categoria = str(row['Categoria']).strip()
                    if not categoria.startswith('  -> ') and not categoria.startswith('  →'):
                        valor = converter_para_float_seguro(row[col])
                        total_col += valor
                col_nome = col.strftime('%m/%y')
                total_secao[col_nome] = formatar_valor_limpo(total_col) if total_col != 0 else '-'
        
        # Cria DataFrame do total
        df_total = pd.DataFrame([{'Categoria': f'🔸 TOTAL {str(titulo)}', 'Tipo': 'TOTAL', **total_secao}])
        
        # Exibe total em destaque
        st.markdown("**TOTAL:**")
        st.dataframe(
            df_total,
            use_container_width=True,
            height=80,
            column_config={
                "Categoria": st.column_config.TextColumn("Total", width="medium")
            }
        )
    
    st.markdown("---")

def criar_tabela_com_total(dados, titulo, colunas_data, cor_emoji=""):
    """Cria tabela com subcategorias e total destacado"""
    if not dados:
        return
        
    st.markdown(f"### {cor_emoji} **{titulo}**")
    
    # Processa dados da seção
    df_secao = pd.DataFrame(dados)
    
    if not df_secao.empty:
        # Formata valores
        df_display = df_secao.copy()
        for col in colunas_data:
            df_display[col] = df_display[col].apply(formatar_valor_limpo)
        
        # Formata colunas de data
        df_display.columns = [col.strftime('%m/%y') if isinstance(col, datetime) else col for col in df_display.columns]
        
        # Exibe tabela principal
        st.dataframe(
            df_display,
            use_container_width=True,
            height=min(len(df_secao) * 35 + 50, 300),
            column_config={
                "Categoria": st.column_config.TextColumn("Categoria", width="medium"),
                "Tipo": st.column_config.TextColumn("Tipo", width="small")
            }
        )
        
        # Calcula e exibe total
        total_secao = {}
        for col in colunas_data:
            total_col = 0
            for _, row in df_secao.iterrows():
                valor = converter_para_float_seguro(row[col])
                total_col += valor
            col_nome = col.strftime('%m/%y')
            total_secao[col_nome] = formatar_valor_limpo(total_col) if total_col != 0 else '-'
        
        # Cria DataFrame do total
        df_total = pd.DataFrame([{'Categoria': f'🔸 TOTAL {str(titulo)}', 'Tipo': 'TOTAL', **total_secao}])
        
        # Exibe total em destaque
        st.markdown("**TOTAL:**")
        st.dataframe(
            df_total,
            use_container_width=True,
            height=80,
            column_config={
                "Categoria": st.column_config.TextColumn("Total", width="medium")
            }
        )
    
    st.markdown("---")

def formatar_valor_visual(valor):
    """Formata valores de forma mais visual com K/M e simplifica percentuais"""
    if valor == 0:
        return '-'
    
    if abs(valor) >= 1000000:
        return f"{valor/1000000:.1f}M"
    elif abs(valor) >= 1000:
        return f"{valor/1000:.0f}K"
    else:
        return f"{valor:.0f}"

def formatar_valor_compacto(valor):
    """Formata valores monetários de forma compacta com R$ para métricas"""
    if valor == 0:
        return 'R$ 0'
    
    if abs(valor) >= 1000000000:
        return f"R$ {valor/1000000000:.1f}B"
    elif abs(valor) >= 1000000:
        return f"R$ {valor/1000000:.1f}M"
    elif abs(valor) >= 100000:
        return f"R$ {valor/1000:.0f}K"
    elif abs(valor) >= 1000:
        return f"R$ {valor/1000:.1f}K"
    else:
        return f"R$ {valor:,.0f}"

def carregar_despesas_detalhadas():
    """Carrega dados detalhados de despesas do arquivo Excel"""
    try:
        # Tenta carregar o arquivo de despesas
        df = pd.read_excel('Despesas.xlsx', header=1)
        return df
    except:
        # Se não conseguir, retorna dados de exemplo
        return pd.DataFrame()

def mostrar_tabela_unificada(df_proj):
    """Mostra uma tabela unificada com todas as projeções de forma clara e apresentável usando dados reais"""
    
    st.subheader("Tabela Unificada de Projeções (Jun/2025 - Dez/2030)")
    
    # Legenda de formatação
    st.info("**Formatação:** K = mil | M = milhão | % simplificado | - = zero")
    
    try:
        # Carrega dados reais da matriz detalhada diretamente
        df_matriz = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
        
        # Identifica colunas de data
        colunas_data = []
        for col in df_matriz.columns:
            if isinstance(col, (pd.Timestamp, datetime)):
                colunas_data.append(col)
        
        # Opção para mostrar/ocultar detalhes
        mostrar_detalhes = st.checkbox("Mostrar detalhes das despesas", value=False, key="mostrar_detalhes_despesas")
        
        # Prepara dados consolidados por ano
        dados_anuais = {}
        
        # Inicializa anos de 2025 a 2030
        for ano in range(2025, 2031):
            dados_anuais[ano] = {
                'receitas_total': 0,
                'custos_total': 0,
                'despesas_total': 0,
                'impostos_federais_total': 0,
                'receitas_detalhes': [],
                'custos_detalhes': [],
                'despesas_detalhes': [],
                'impostos_detalhes': []
            }
        
        # USA TOTAIS DIRETOS DA PLANILHA (abordagem mais confiável)
        for col in colunas_data:
            if isinstance(col, (datetime, pd.Timestamp)):
                ano = col.year
                if ano in dados_anuais:
                    
                    # RECEITAS - usa TOTAL RECEITAS direto
                    receitas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL RECEITAS']
                    if not receitas_linha.empty:
                        valor_receita = converter_para_float_seguro(receitas_linha[col].iloc[0])
                        dados_anuais[ano]['receitas_total'] += valor_receita
                    
                    # CUSTOS - usa TOTAL CUSTOS direto
                    custos_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL CUSTOS']
                    if not custos_linha.empty:
                        valor_custo = converter_para_float_seguro(custos_linha[col].iloc[0])
                        dados_anuais[ano]['custos_total'] += valor_custo
                    
                    # DESPESAS - usa TOTAL DESPESAS direto (sem duplicação!)
                    despesas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL DESPESAS']
                    if not despesas_linha.empty:
                        valor_despesa = converter_para_float_seguro(despesas_linha[col].iloc[0])
                        dados_anuais[ano]['despesas_total'] += valor_despesa
                        
                        # Para detalhamento, pega as 8 categorias pai de despesas
                        categorias_despesas = ['Marketing e Comercial', 'Produto', 'Impostos', 'Juridico', 'Pro labore', 'Serviços', 'Equipe', 'Sistemas']
                        for cat_despesa in categorias_despesas:
                            cat_linha = df_matriz[df_matriz['Categoria'] == cat_despesa]
                            if not cat_linha.empty:
                                valor_cat = converter_para_float_seguro(cat_linha[col].iloc[0])
                                if valor_cat > 0:
                                    dados_anuais[ano]['despesas_detalhes'].append({
                                        'item': cat_despesa,
                                        'valor': valor_cat
                                    })
        
        # Prepara os dados para a tabela unificada
        tabela_unificada = []
    
        # Adiciona linhas de receitas
        tabela_unificada.append({
            'Categoria': 'RECEITAS',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # Receitas brutas (dados reais)
        linha_receitas = {'Categoria': '   Receitas Brutas'}
        total_receitas = 0
        for ano in range(2025, 2031):
            valor = dados_anuais[ano]['receitas_total']
            linha_receitas[str(ano)] = formatar_valor_visual(valor)
            total_receitas += valor
        linha_receitas['Total Período'] = formatar_valor_visual(total_receitas)
        tabela_unificada.append(linha_receitas)
    
        # Linha vazia
        tabela_unificada.append({
            'Categoria': '',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # Custos e Impostos sobre Vendas
        tabela_unificada.append({
            'Categoria': 'CUSTOS E IMPOSTOS S/ VENDAS',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # Calcular impostos MES A MES (igual ao Por Ano Detalhado)
        impostos_por_ano = {}
        for ano in range(2025, 2031):
            impostos_por_ano[ano] = {
                'ISS': 0,
                'PIS': 0,
                'COFINS': 0,
                'IRPJ': 0,
                'CSLL': 0
            }
        
        # Calcula impostos MÊS A MÊS (abordagem correta)
        for col in colunas_data:
            if isinstance(col, (datetime, pd.Timestamp)):
                ano = col.year
                if ano in impostos_por_ano and col >= pd.to_datetime('2025-10-01'):
                    
                    # Pega receita do mês
                    receitas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL RECEITAS']
                    if not receitas_linha.empty:
                        receita_mes = converter_para_float_seguro(receitas_linha[col].iloc[0])
                        
                        if receita_mes > 0:
                            # Divisão: 40% serviços digitais, 60% venda de livros
                            receitas_servicos = receita_mes * 0.4
                            receitas_livros = receita_mes * 0.6
                            
                            # USA A FUNÇÃO CENTRALIZADA PARA GARANTIR CONSISTÊNCIA
                            impostos_mes_dict = calcular_impostos_brasileiros(receitas_servicos, receitas_livros, col)
                            
                            # Adiciona aos totais anuais
                            for tributo, valor in impostos_mes_dict.items():
                                if tributo in impostos_por_ano[ano]:
                                    impostos_por_ano[ano][tributo] += valor
        
        # Linha de custos operacionais (usando total direto da planilha)
        linha_custos_op = {'Categoria': '   Custos Operacionais'}
        total_custos_op = 0
        for ano in range(2025, 2031):
            valor_custos = dados_anuais[ano]['custos_total']  # Já é o total correto da planilha
            linha_custos_op[str(ano)] = formatar_valor_visual(valor_custos)
            total_custos_op += valor_custos
        linha_custos_op['Total Período'] = formatar_valor_visual(total_custos_op)
        tabela_unificada.append(linha_custos_op)
        
        # Linha ISS + PIS + COFINS
        linha_impostos_venda = {'Categoria': '   ISS + PIS + COFINS'}
        total_impostos_venda = 0
        for ano in range(2025, 2031):
            valor_impostos = impostos_por_ano[ano]['ISS'] + impostos_por_ano[ano]['PIS'] + impostos_por_ano[ano]['COFINS']
            linha_impostos_venda[str(ano)] = formatar_valor_visual(valor_impostos)
            total_impostos_venda += valor_impostos
        linha_impostos_venda['Total Período'] = formatar_valor_visual(total_impostos_venda)
        tabela_unificada.append(linha_impostos_venda)
    
        # Linha vazia
        tabela_unificada.append({
            'Categoria': '',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # Despesas (dados reais)
        tabela_unificada.append({
            'Categoria': 'DESPESAS',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # Mostrar detalhes das despesas se solicitado
        if mostrar_detalhes:
            # Agrupa despesas por categoria para visualização detalhada
            categorias_despesas = {}
            for ano in range(2025, 2031):
                for item in dados_anuais[ano]['despesas_detalhes']:
                    categoria_item = item['item']
                    if categoria_item not in categorias_despesas:
                        categorias_despesas[categoria_item] = {}
                        for ano_init in range(2025, 2031):
                            categorias_despesas[categoria_item][str(ano_init)] = 0
                    categorias_despesas[categoria_item][str(ano)] += item['valor']
            
            # Adiciona linhas detalhadas para cada categoria de despesa
            for categoria, valores_anuais in categorias_despesas.items():
                linha_categoria = {'Categoria': f'   {categoria}'}
                total_categoria = 0
                for ano in range(2025, 2031):
                    valor = valores_anuais[str(ano)]
                    linha_categoria[str(ano)] = formatar_valor_visual(valor) if valor > 0 else '-'
                    total_categoria += valor
                linha_categoria['Total Período'] = formatar_valor_visual(total_categoria)
                tabela_unificada.append(linha_categoria)
        
        # Total de despesas
        linha_total_desp = {'Categoria': '   TOTAL DESPESAS'}
        total_desp_geral = 0
        for ano in range(2025, 2031):
            valor = dados_anuais[ano]['despesas_total']
            linha_total_desp[str(ano)] = formatar_valor_visual(valor)
            total_desp_geral += valor
        linha_total_desp['Total Período'] = formatar_valor_visual(total_desp_geral)
        tabela_unificada.append(linha_total_desp)
        
        # Linha vazia
        tabela_unificada.append({
            'Categoria': '',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # Impostos Federais sobre Lucro
        tabela_unificada.append({
            'Categoria': 'IMPOSTOS FEDERAIS S/ LUCRO',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # IRPJ
        linha_irpj = {'Categoria': '   IRPJ (15% + adicional)'}
        total_irpj = 0
        for ano in range(2025, 2031):
            valor = impostos_por_ano[ano]['IRPJ']
            linha_irpj[str(ano)] = formatar_valor_visual(valor)
            total_irpj += valor
        linha_irpj['Total Período'] = formatar_valor_visual(total_irpj)
        tabela_unificada.append(linha_irpj)
        
        # CSLL
        linha_csll = {'Categoria': '   CSLL (9%)'}
        total_csll = 0
        for ano in range(2025, 2031):
            valor = impostos_por_ano[ano]['CSLL']
            linha_csll[str(ano)] = formatar_valor_visual(valor)
            total_csll += valor
        linha_csll['Total Período'] = formatar_valor_visual(total_csll)
        tabela_unificada.append(linha_csll)
        
        # Separador
        tabela_unificada.append({
            'Categoria': '═' * 30,
            '2025': '═' * 15,
            '2026': '═' * 15,
            '2027': '═' * 15,
            '2028': '═' * 15,
            '2029': '═' * 15,
            '2030': '═' * 15,
            'Total Período': '═' * 15
        })
        
        # Resultados (calculados)
        tabela_unificada.append({
            'Categoria': 'RESULTADOS',
            '2025': '',
            '2026': '',
            '2027': '',
            '2028': '',
            '2029': '',
            '2030': '',
            'Total Período': ''
        })
        
        # Lucro Bruto (Receitas - Custos - ISS/PIS/COFINS)
        linha_lucro_bruto = {'Categoria': '   Lucro Bruto'}
        total_lucro_bruto = 0
        for ano in range(2025, 2031):
            receitas = dados_anuais[ano]['receitas_total']
            # USA TOTAL CUSTOS DIRETO DA PLANILHA
            custos = dados_anuais[ano]['custos_total']
            # Impostos sobre vendas
            impostos_vendas = impostos_por_ano[ano]['ISS'] + impostos_por_ano[ano]['PIS'] + impostos_por_ano[ano]['COFINS']
            valor = receitas - custos - impostos_vendas
            linha_lucro_bruto[str(ano)] = formatar_valor_visual(valor)
            total_lucro_bruto += valor
        linha_lucro_bruto['Total Período'] = formatar_valor_visual(total_lucro_bruto)
        tabela_unificada.append(linha_lucro_bruto)
        
        # Lucro Operacional (Lucro Bruto - Despesas)
        linha_lucro_op = {'Categoria': '   Lucro Operacional'}
        total_lucro_op = 0
        for ano in range(2025, 2031):
            receitas = dados_anuais[ano]['receitas_total']
            # USA TOTAL CUSTOS DIRETO DA PLANILHA
            custos = dados_anuais[ano]['custos_total']
            # Impostos sobre vendas
            impostos_vendas = impostos_por_ano[ano]['ISS'] + impostos_por_ano[ano]['PIS'] + impostos_por_ano[ano]['COFINS']
            despesas = dados_anuais[ano]['despesas_total']
            valor = receitas - custos - impostos_vendas - despesas
            linha_lucro_op[str(ano)] = formatar_valor_visual(valor)
            total_lucro_op += valor
        linha_lucro_op['Total Período'] = formatar_valor_visual(total_lucro_op)
        tabela_unificada.append(linha_lucro_op)
        
        # Resultado Líquido Final
        linha_resultado = {'Categoria': 'RESULTADO LÍQUIDO'}
        total_resultado = 0
        for ano in range(2025, 2031):
            receitas = dados_anuais[ano]['receitas_total']
            # USA TOTAL CUSTOS DIRETO DA PLANILHA (mesma lógica das outras funções)
            custos = dados_anuais[ano]['custos_total']
            # Todos os impostos
            impostos_total = (impostos_por_ano[ano]['ISS'] + impostos_por_ano[ano]['PIS'] + 
                             impostos_por_ano[ano]['COFINS'] + impostos_por_ano[ano]['IRPJ'] + 
                             impostos_por_ano[ano]['CSLL'])
            despesas = dados_anuais[ano]['despesas_total']
            valor = receitas - custos - impostos_total - despesas
            linha_resultado[str(ano)] = formatar_valor_visual(valor)
            total_resultado += valor
        linha_resultado['Total Período'] = formatar_valor_visual(total_resultado)
        tabela_unificada.append(linha_resultado)
        
        # Converte para DataFrame
        df_tabela = pd.DataFrame(tabela_unificada)
        
        # Exibe a tabela
        st.dataframe(
            df_tabela,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Categoria": st.column_config.TextColumn("Categoria", width="large"),
                "2025": st.column_config.TextColumn("2025", width="medium"),
                "2026": st.column_config.TextColumn("2026", width="medium"),
                "2027": st.column_config.TextColumn("2027", width="medium"),
                "2028": st.column_config.TextColumn("2028", width="medium"),
                "2029": st.column_config.TextColumn("2029", width="medium"),
                "2030": st.column_config.TextColumn("2030", width="medium"),
                "Total Período": st.column_config.TextColumn("Total Período", width="medium"),
            }
        )
        
        # Resumo dos Impostos com detalhamento
        st.markdown("### **Detalhamento Tributário - Lucro Presumido (40% Serviços Digitais + 60% Venda de Livros)**")
        
        if st.button("Mostrar/Ocultar Detalhamento Tributário", key="toggle_tributario"):
            if "show_tributario" not in st.session_state:
                st.session_state.show_tributario = False
            st.session_state.show_tributario = not st.session_state.show_tributario
        
        if st.session_state.get("show_tributario", False):
            st.markdown("### Resumo dos Impostos por Ano")
            
            # Tabela de impostos detalhada
            impostos_detalhados = []
            for ano in range(2025, 2031):
                receitas_ano = dados_anuais[ano]['receitas_total']
                if receitas_ano > 0:
                    impostos_detalhados.append({
                        'Ano': str(ano),
                        'Receita Total': formatar_moeda_br(receitas_ano),
                        'ISS (2% s/ serviços)': formatar_moeda_br(impostos_por_ano[ano]['ISS']),
                        'PIS (0,65%)': formatar_moeda_br(impostos_por_ano[ano]['PIS']),
                        'COFINS (3%)': formatar_moeda_br(impostos_por_ano[ano]['COFINS']),
                        'IRPJ': formatar_moeda_br(impostos_por_ano[ano]['IRPJ']),
                        'CSLL (9%)': formatar_moeda_br(impostos_por_ano[ano]['CSLL']),
                        'Total Impostos': formatar_moeda_br(
                            impostos_por_ano[ano]['ISS'] + impostos_por_ano[ano]['PIS'] + 
                            impostos_por_ano[ano]['COFINS'] + impostos_por_ano[ano]['IRPJ'] + 
                            impostos_por_ano[ano]['CSLL']
                        ),
                        'Carga Tributária': f"{((impostos_por_ano[ano]['ISS'] + impostos_por_ano[ano]['PIS'] + impostos_por_ano[ano]['COFINS'] + impostos_por_ano[ano]['IRPJ'] + impostos_por_ano[ano]['CSLL']) / receitas_ano * 100):.1f}%"
                    })
            
            if impostos_detalhados:
                df_impostos = pd.DataFrame(impostos_detalhados)
                st.dataframe(df_impostos, use_container_width=True, hide_index=True)
                
                # Explicação das alíquotas
                st.markdown("""
                #### 📄 **Base de Cálculo - Lucro Presumido:**
                - **Serviços Digitais (40% da receita):**
                  - ISS: 2% sobre o faturamento de serviços
                  - Base IRPJ/CSLL: 32% de presunção sobre a receita
                - **Venda de Livros (60% da receita):**
                  - ISS: Isento (não incide sobre venda de produtos)
                  - Base IRPJ/CSLL: 1,6% de presunção sobre a receita (benefício fiscal)
                - **PIS/COFINS:** Incidem sobre o faturamento total (0,65% + 3%)
                - **IRPJ:** 15% sobre a base + adicional de 10% sobre o que exceder R$ 20.000/mês
                - **CSLL:** 9% sobre a base de cálculo
                """)
        
        # Seção de detalhamento com dados reais
        st.markdown("### **Ver Detalhamento das Despesas por Item**")
        
        if st.button("Mostrar/Ocultar Detalhamento das Despesas", key="toggle_despesas"):
            if "show_despesas" not in st.session_state:
                st.session_state.show_despesas = False
            st.session_state.show_despesas = not st.session_state.show_despesas
        
        if st.session_state.get("show_despesas", False):
            st.markdown("### Despesas Detalhadas da Matriz Financeira")
            
            for ano in range(2025, 2031):
                if dados_anuais[ano]['despesas_detalhes']:
                    st.markdown(f"#### Despesas de {ano}")
                    
                    if st.button(f"Mostrar/Ocultar Despesas de {ano}", key=f"toggle_ano_{ano}"):
                        if f"show_ano_{ano}" not in st.session_state:
                            st.session_state[f"show_ano_{ano}"] = False
                        st.session_state[f"show_ano_{ano}"] = not st.session_state[f"show_ano_{ano}"]
                    
                    if st.session_state.get(f"show_ano_{ano}", False):
                        despesas_ano = dados_anuais[ano]['despesas_detalhes']
                        
                        col1, col2 = st.columns([2, 1])
                        
                        with col1:
                            # Tabela de despesas
                            df_despesas_ano = pd.DataFrame([
                                {'Item': item['item'], 'Valor': formatar_moeda_br(item['valor'])}
                                for item in despesas_ano
                            ])
                            st.dataframe(df_despesas_ano, use_container_width=True, hide_index=True)
                        
                        with col2:
                            # Gráfico de pizza das despesas
                            if len(despesas_ano) > 1:
                                labels = [item['item'][:20] + '...' if len(item['item']) > 20 else item['item'] for item in despesas_ano]
                                valores = [item['valor'] for item in despesas_ano]
                                
                                fig = go.Figure(data=[go.Pie(
                                    labels=labels,
                                    values=valores,
                                    hole=.3,
                                    marker_colors=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
                                                 '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9']
                                )])
                                
                                fig.update_layout(
                                    showlegend=True,
                                    height=300,
                                    margin=dict(l=0, r=0, t=30, b=0),
                                    title=f"Distribuição {ano}",
                                    title_x=0.5,
                                    legend=dict(
                                        orientation="v",
                                        yanchor="middle",
                                        y=0.5,
                                        xanchor="left",
                                        x=1.02,
                                        font=dict(size=10)
                                    )
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                        
                        total_ano = sum(item['valor'] for item in despesas_ano)
                        st.info(f"**Total {ano}:** {formatar_moeda_br(total_ano)}")
        
    except Exception as e:
        st.error(f"Erro ao carregar dados da matriz: {e}")
        st.info("💡 Verifique se o arquivo 'Matriz financeira.xlsx' está disponível.")
        return

def mostrar_matriz_detalhada():
    """Mostra a matriz detalhada das projeções com visual aprimorado e organizado"""
    
    st.subheader("Matriz Detalhada de Projeções")
    
    try:
        # Processa dados organizados
        secoes, colunas_data = processar_dados_organizados()
        
        # Header informativo
        col1, col2, col3 = st.columns([2, 1, 1])
        with col1:
            st.info("**Período:** Jun/2025 → Dez/2030 (67 meses)")
            st.caption("**Formato:** k = mil | M = milhão | ✓ = positivo | ✗ = negativo")
            st.warning("🏛️ **Impostos:** Lucro Presumido a partir Out/2025 (40% serviços + 60% livros)")
        
        with col2:
            try:
                # Calcula total de receitas
                total_receitas = 0
                if secoes['RECEITAS']:
                    for row in secoes['RECEITAS']:
                        for col in colunas_data:
                            valor_receita = converter_para_float_seguro(row[col])
                            total_receitas += valor_receita
                st.metric("Total Receitas", formatar_valor_compacto(total_receitas))
            except:
                st.metric("Total Receitas", "Calculando...")
        
        with col3:
            try:
                # Calcula resultado final USANDO VALOR DIRETO DA PLANILHA
                df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
                linha_resultado_original = df_matriz_original[df_matriz_original['Categoria'] == 'RESULTADO LÍQUIDO']
                
                if not linha_resultado_original.empty:
                    total_resultado = 0
                    for col in colunas_data:
                        valor = linha_resultado_original[col].iloc[0]
                        if pd.notna(valor):
                            total_resultado += float(valor)
                    cor = "normal" if total_resultado >= 0 else "inverse"
                    st.metric("Resultado Final", formatar_valor_compacto(total_resultado), delta_color=cor)
                else:
                    st.metric("Resultado Final", "Calculando...")
            except:
                st.metric("Resultado Final", "Calculando...")
        
        st.markdown("---")
        
        # Exibe seções organizadas com totais
        criar_tabela_com_total(secoes['RECEITAS'], "RECEITAS", colunas_data, "")
        criar_tabela_com_total(secoes['CUSTOS'], "CUSTOS (+ ISS, PIS, COFINS)", colunas_data, "")
        criar_tabela_com_total(secoes['IMPOSTOS_FEDERAIS'], "IMPOSTOS FEDERAIS (IRPJ, CSLL)", colunas_data, "🏛️")
        criar_tabela_despesas_hierarquica(secoes['DESPESAS'], "DESPESAS", colunas_data, "")
        
        # RESULTADO LÍQUIDO (sem total pois já é o resultado final)
        if secoes['RESULTADO']:
            st.markdown("### **RESULTADO LÍQUIDO**")
            df_resultado = pd.DataFrame(secoes['RESULTADO'])
            df_resultado_display = df_resultado.copy()
            
            for col in colunas_data:
                df_resultado_display[col] = df_resultado_display[col].apply(formatar_resultado_limpo)
            
            df_resultado_display.columns = [col.strftime('%m/%y') if isinstance(col, datetime) else col for col in df_resultado_display.columns]
            
            st.dataframe(
                df_resultado_display,
                use_container_width=True,
                height=100,
                column_config={
                    "Categoria": st.column_config.TextColumn("📊 Categoria", width="medium")
                }
            )
        
        # LUCRO LÍQUIDO E MARGEM LÍQUIDA
        st.markdown("---")
        st.markdown("### **LUCRO LÍQUIDO E MARGEM**")
        
        # Calcula totais por mês USANDO VALORES DIRETOS DA PLANILHA
        lucro_liquido_mensal = {}
        margem_liquida_mensal = {}
        
        # Carrega valores diretos da planilha para totais
        df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
        linha_total_receitas = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL RECEITAS']
        linha_total_custos = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL CUSTOS']
        linha_total_despesas = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
        linha_resultado_liquido = df_matriz_original[df_matriz_original['Categoria'] == 'RESULTADO LÍQUIDO']
        
        for col in colunas_data:
            # Usa valores DIRETOS da planilha
            receitas_mes = float(linha_total_receitas[col].iloc[0]) if not linha_total_receitas.empty and pd.notna(linha_total_receitas[col].iloc[0]) else 0
            custos_mes = float(linha_total_custos[col].iloc[0]) if not linha_total_custos.empty and pd.notna(linha_total_custos[col].iloc[0]) else 0
            despesas_mes = float(linha_total_despesas[col].iloc[0]) if not linha_total_despesas.empty and pd.notna(linha_total_despesas[col].iloc[0]) else 0
            resultado_mes = float(linha_resultado_liquido[col].iloc[0]) if not linha_resultado_liquido.empty and pd.notna(linha_resultado_liquido[col].iloc[0]) else 0
            
            # Usa o resultado direto da planilha
            lucro_liquido_mensal[col] = resultado_mes
            
            # Calcula margem líquida
            if receitas_mes > 0:
                margem_mes = (resultado_mes / receitas_mes) * 100
            else:
                margem_mes = 0
            margem_liquida_mensal[col] = margem_mes
        
        # Cria DataFrames para exibição
        df_lucro = pd.DataFrame([{
            'Categoria': 'LUCRO LÍQUIDO',
            **{col.strftime('%m/%y'): formatar_valor_limpo(lucro_liquido_mensal[col]) for col in colunas_data}
        }])
        
        df_margem = pd.DataFrame([{
            'Categoria': 'MARGEM LÍQUIDA (%)',
            **{col.strftime('%m/%y'): f"{margem_liquida_mensal[col]:.1f}%" if margem_liquida_mensal[col] != 0 else "-" for col in colunas_data}
        }])
        
        # Exibe Lucro Líquido
        st.dataframe(
            df_lucro,
            use_container_width=True,
            height=80,
            column_config={
                "Categoria": st.column_config.TextColumn("📊 Resultado", width="medium")
            }
        )
        
        # Exibe Margem Líquida
        st.dataframe(
            df_margem,
            use_container_width=True,
            height=80,
            column_config={
                "Categoria": st.column_config.TextColumn("📊 Indicador", width="medium")
            }
        )
        
        
            
    except Exception as e:
        st.error(f"❌ Erro ao carregar matriz detalhada: {str(e)}")
        st.info("💡 Certifique-se de que o arquivo 'Matriz financeira.xlsx' existe na pasta do projeto")

def mostrar_consolidado_5_anos(df_proj):
    """Mostra visão consolidada de 5 anos"""
    
    st.subheader("Visão Consolidada - 5 Anos")
    
    # Gráfico de evolução anual
    st.subheader("Evolução Anual")
    
    fig = go.Figure()
    
    # Receitas
    fig.add_trace(go.Scatter(
        x=df_proj['Ano'],
        y=df_proj['Receitas_Brutas'],
        mode='lines+markers',
        name='Receitas',
        line=dict(color='#28a745', width=3),
        marker=dict(size=8)
    ))
    
    # Despesas
    fig.add_trace(go.Scatter(
        x=df_proj['Ano'],
        y=df_proj['Despesas_Totais'],
        mode='lines+markers',
        name='Despesas',
        line=dict(color='#dc3545', width=3),
        marker=dict(size=8)
    ))
    
    # Resultado
    fig.add_trace(go.Scatter(
        x=df_proj['Ano'],
        y=df_proj['Resultado_Liquido'],
        mode='lines+markers',
        name='Resultado Líquido',
        line=dict(color='#2a5298', width=3),
        marker=dict(size=8)
    ))
    
    fig.update_layout(
        title="Evolução Financeira - 5 Anos",
        xaxis_title="Ano",
        yaxis_title="Valores",
        yaxis_tickformat='.2s',  # Formato científico simplificado (1M, 2.5K, etc)
        height=500,
        hovermode='x unified'
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Tabela consolidada
    st.subheader("Tabela Consolidada")
    
    # Legenda de formatação
    st.info("**Formatação:** K = mil | M = milhão | - = zero")
    
    # Formatar valores para exibição visual
    df_display = df_proj.copy()
    for col in ['Receitas_Brutas', 'Custos_Operacionais', 'Despesas_Totais', 'Impostos_Totais', 'Resultado_Liquido']:
        if col in df_display.columns:
            df_display[col] = df_display[col].apply(formatar_valor_visual)
    
    # Formatar margem líquida como percentual sem casas decimais
    if 'Margem_Liquida' in df_display.columns:
        df_display['Margem_Liquida'] = df_display['Margem_Liquida'].apply(lambda x: f"{x:.0f}%" if x != 0 else "-")
    
    # Renomear colunas
    df_display = df_display.rename(columns={
        'Receitas_Brutas': 'Receitas',
        'Custos_Operacionais': 'Custos',
        'Despesas_Totais': 'Despesas',
        'Impostos_Totais': 'Impostos',
        'Resultado_Liquido': 'Resultado',
        'Margem_Liquida': 'Margem %'
    })
    
    st.dataframe(df_display, use_container_width=True, hide_index=True)
    

def exportar_projecoes_excel(df_proj):
    """Exporta projeções de 5 anos para Excel"""
    
    try:
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Aba consolidada
            df_proj.to_excel(writer, sheet_name='Projecoes_5_Anos', index=False)
            
            # Aba com análises
            analise = []
            analise.append(['ANÁLISE DE CRESCIMENTO - 5 ANOS', '', '', '', ''])
            analise.append(['', '', '', '', ''])
            
            for i in range(1, len(df_proj)):
                ano_atual = df_proj.iloc[i]['Ano']
                ano_anterior = df_proj.iloc[i-1]['Ano']
                
                cresc_receitas = ((df_proj.iloc[i]['Receitas_Brutas'] / df_proj.iloc[i-1]['Receitas_Brutas']) - 1) * 100
                cresc_resultado = ((df_proj.iloc[i]['Resultado_Liquido'] / df_proj.iloc[i-1]['Resultado_Liquido']) - 1) * 100 if df_proj.iloc[i-1]['Resultado_Liquido'] != 0 else 0
                
                analise.append([
                    f'Crescimento {ano_anterior}-{ano_atual}',
                    f'{cresc_receitas:.1f}%',
                    f'{cresc_resultado:.1f}%',
                    '',
                    ''
                ])
            
            df_analise = pd.DataFrame(analise, columns=['Indicador', 'Receitas', 'Resultado', 'Obs1', 'Obs2'])
            df_analise.to_excel(writer, sheet_name='Analise_Crescimento', index=False)
        
        output.seek(0)
        st.download_button(
            label="💾 Baixar Projeções Excel",
            data=output,
            file_name=f"projecoes_5_anos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.success("✅ Arquivo de projeções pronto para download!")
        
    except Exception as e:
        st.error(f"❌ Erro ao gerar arquivo de projeções: {e}")

def sistema_unificado():
    """Sistema Unificado - Exporta matriz financeira no mesmo formato da seção Matriz Financeira"""
    
    st.title("💎 Sistema Unificado")
    st.markdown("---")
    
    st.markdown("""
    ### **Exportação da Matriz Financeira Unificada**
    
    Este sistema usa o **mesmo formato** da seção "Matriz Financeira" para exportar 
    sua planilha das **Projeções 5 Anos** de forma organizada e bonitinha.
    
    **📥 Características do Export:**
    - 🎨 **Formato igual** ao da Matriz Financeira existente
    - 📊 **Dados das Projeções 5 Anos** organizados
    - 💼 **Planilha única** unificada e editada
    - ✨ **Layout limpo** e profissional
    """)
    
    # Botão principal de exportação
    st.markdown("### **EXPORTAR MATRIZ DAS PROJEÇÕES**")
    
    if st.button("📥 **EXPORTAR MATRIZ FINANCEIRA PARA EXCEL**", type="primary", use_container_width=True):
        try:
            with st.spinner("Gerando matriz financeira unificada..."):
                
                # Usa os mesmos dados das projeções 5 anos
                arquivo_export = "matriz_financeira_projecoes.xlsx"
                
                # Carrega dados das projeções
                try:
                    df_matriz = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
                    
                    with pd.ExcelWriter(arquivo_export, engine='openpyxl') as writer:
                        
                        # Exporta exatamente como está na Matriz Detalhada - formato unificado
                        df_matriz.to_excel(writer, sheet_name='Matriz_Unificada', index=False)
                        
                        # Aplica formatação profissional
                        workbook = writer.book
                        worksheet = writer.sheets['Matriz_Unificada']
                        
                        # Formatação das células
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        
                        # Estilo para header
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Estilo para categorias principais
                        categoria_font = Font(bold=True)
                        categoria_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        
                        # Aplica formatação ao header (primeira linha)
                        for col_num in range(1, df_matriz.shape[1] + 1):
                            cell = worksheet.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Identifica e formata categorias principais
                        categorias_principais = ['RECEITAS', 'CUSTOS', 'DESPESAS', 'IMPOSTOS', 
                                               'TOTAL RECEITAS', 'TOTAL CUSTOS', 'TOTAL DESPESAS', 
                                               'RESULTADO LÍQUIDO']
                        
                        for row_num in range(2, df_matriz.shape[0] + 2):
                            categoria_cell = worksheet.cell(row=row_num, column=1)
                            if categoria_cell.value in categorias_principais:
                                # Formata toda a linha da categoria principal
                                for col_num in range(1, df_matriz.shape[1] + 1):
                                    cell = worksheet.cell(row=row_num, column=col_num)
                                    cell.font = categoria_font
                                    cell.fill = categoria_fill
                        
                        # Ajusta largura das colunas
                        from openpyxl.utils import get_column_letter
                        
                        worksheet.column_dimensions['A'].width = 25
                        worksheet.column_dimensions['B'].width = 15
                        
                        # Usa get_column_letter para suportar muitas colunas (AA, AB, etc.)
                        for col_num in range(3, df_matriz.shape[1] + 1):
                            col_letter = get_column_letter(col_num)
                            worksheet.column_dimensions[col_letter].width = 12
                    
                    st.success(f"✅ **Matriz exportada:** {arquivo_export}")
                    
                    # Botão para download
                    with open(arquivo_export, 'rb') as f:
                        st.download_button(
                            label="💾 **BAIXAR ARQUIVO EXCEL**",
                            data=f.read(),
                            file_name=arquivo_export,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="secondary"
                        )
                        
                except Exception as e:
                    st.error(f"❌ Erro ao carregar Matriz financeira.xlsx: {e}")
                    
        except Exception as e:
            st.error(f"❌ Erro na exportação: {e}")
    
    # Informações sobre o arquivo
    st.markdown("---")
    st.markdown("""
    ### **Informações do Sistema**
    
    **Arquivo fonte:** `Matriz financeira.xlsx` → Aba "Matriz Detalhada"
    
    **📥 Arquivo exportado:** `matriz_financeira_projecoes.xlsx`
    
    **🎨 Formatação aplicada:**
    - Headers com fundo azul e texto branco
    - Categorias principais destacadas em azul claro
    - Colunas ajustadas para melhor visualização
    - Layout profissional e organizado
    
    **✨ Vantagens:**
    - Mesma formatação da seção "Matriz Financeira" 
    - Dados unificados em planilha única
    - Pronto para edição e apresentação
    - Export direto para Excel formatado
    """)


def editor_categorias(df_despesas):
    """Interface para editar categorias das despesas"""
    
    st.header("✏️ Editor de Categorias")
    
    st.markdown("""
    <div class="edit-section">
        <h4>🛠️ Ferramenta de Edição</h4>
        <p>Aqui você pode reclassificar despesas e ajustar categorias conforme necessário.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Seletor de categoria atual
    categorias_atuais = ['Todas'] + sorted(df_despesas['Categoria'].unique().tolist())
    categoria_filtro = st.selectbox("🏷️ Filtrar por categoria atual:", categorias_atuais)
    
    # Filtra dados
    if categoria_filtro != 'Todas':
        df_filtrado = df_despesas[df_despesas['Categoria'] == categoria_filtro].copy()
    else:
        df_filtrado = df_despesas.copy()
    
    # Mostra itens para edição
    if not df_filtrado.empty:
        st.subheader(f"📝 Itens da categoria: {categoria_filtro}")
        
        # Lista para edição
        for idx, row in df_filtrado.iterrows():
            with st.expander(f"{row['Descricao']} - {formatar_moeda_br(row['Valor'])}", expanded=False):
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write(f"**Categoria atual:** {row['Categoria']}")
                    st.write(f"**Área:** {row['Area']}")
                
                with col2:
                    st.write(f"**Valor:** {formatar_moeda_br(row['Valor'])}")
                    st.write(f"**Status:** {row['Status']}")
                
                with col3:
                    st.write(f"**Mês:** {row['Mes']}")
                    st.write(f"**Data:** {row['Data'].strftime('%d/%m/%Y')}")
                
                # Opções de edição
                nova_categoria = st.selectbox(
                    "Nova categoria:",
                    ['Sistemas', 'Serviços', 'ADS', 'Juridico', 'Pro labore', 'Marketing', 'Financeiro', 'Impostos'],
                    index=0,
                    key=f"cat_{idx}"
                )
                
                if st.button(f"💾 Salvar alteração", key=f"save_{idx}"):
                    st.success(f"✅ Categoria alterada para: {nova_categoria}")
                    st.info("💡 Em uma versão completa, isso salvaria no banco de dados")

def mostrar_matriz_resumida(df_despesas, df_receitas):
    """Matriz resumida com despesas agregadas por categoria para o Overview"""
    
    st.header("Matriz Financeira - Resumo")
    
    # Cria dados consolidados
    dados_matriz = []
    
    # Header
    meses = ['Jun/25', 'Jul/25', 'Ago/25', 'Set/25']
    
    # RECEITAS
    
    # Receitas por status
    for status in ['Realizado', 'Projetado']:
        linha = [f'Receitas {status}', '']
        total_status = 0
        
        for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
            valor = df_receitas[
                (df_receitas['Data'].dt.strftime('%Y-%m') == mes) & 
                (df_receitas['Status'] == status)
            ]['Valor'].sum()
            linha.append(formatar_moeda_br(valor))
            total_status += valor
        
        linha.append(formatar_moeda_br(total_status))
        dados_matriz.append(linha)
    
    # Total receitas
    linha_total_rec = ['TOTAL RECEITAS', '']
    total_geral_rec = 0
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        valor = df_receitas[df_receitas['Data'].dt.strftime('%Y-%m') == mes]['Valor'].sum()
        linha_total_rec.append(formatar_moeda_br(valor))
        total_geral_rec += valor
    linha_total_rec.append(formatar_moeda_br(total_geral_rec))
    dados_matriz.append(linha_total_rec)
    
    # Separador
    dados_matriz.append(['', '', '', '', '', ''])
    
    # CUSTOS (Tarifas)
    dados_matriz.append(['CUSTOS', '', '', '', '', ''])
    
    # Receitas brutas por mês para cálculo das tarifas
    receitas_por_mes = {
        '2025-06': 198198.40,  # Junho bruto
        '2025-07': 433176.64,  # Julho bruto  
        '2025-08': 146682.25,  # Agosto bruto
        '2025-09': 22278.47    # Setembro bruto (realizado)
    }
    
    # Tarifa de adquirente (2.5% sobre receitas - TODOS os meses)
    linha_adquirente = ['Tarifa Adquirente (2.5%)', '']
    total_adquirente = 0
    
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        if mes in receitas_por_mes:
            tarifa = receitas_por_mes[mes] * 0.025
            linha_adquirente.append(formatar_moeda_br(tarifa))
            total_adquirente += tarifa
        else:
            linha_adquirente.append('-')
    linha_adquirente.append(formatar_moeda_br(total_adquirente))
    dados_matriz.append(linha_adquirente)
    
    # Tarifa de antecipação (3.33% apenas setembro em diante)
    linha_antecipacao = ['Tarifa Antecipação (3.33%)', '']
    total_antecipacao = 0
    
    tarifas_antecipacao = {
        '2025-06': 0,           # Não teve
        '2025-07': 0,           # Não teve
        '2025-08': 0,           # Não teve
        '2025-09': receitas_por_mes['2025-09'] * 0.0333  # Setembro: 3.33% sobre realizado
    }
    
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        tarifa = tarifas_antecipacao.get(mes, 0)
        if tarifa > 0:
            linha_antecipacao.append(formatar_moeda_br(tarifa))
            total_antecipacao += tarifa
        else:
            linha_antecipacao.append('-')
    linha_antecipacao.append(formatar_moeda_br(total_antecipacao))
    dados_matriz.append(linha_antecipacao)
    
    # Total custos (adquirente + antecipação)
    linha_total_custos = ['TOTAL CUSTOS', '']
    total_custos_geral = 0
    
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        # Adquirente todos os meses
        custo_adquirente = receitas_por_mes.get(mes, 0) * 0.025
        # Antecipação só setembro em diante
        custo_antecipacao = tarifas_antecipacao.get(mes, 0)
        
        total_mes = custo_adquirente + custo_antecipacao
        linha_total_custos.append(formatar_moeda_br(total_mes))
        total_custos_geral += total_mes
    
    linha_total_custos.append(formatar_moeda_br(total_custos_geral))
    dados_matriz.append(linha_total_custos)
    
    # Separador
    dados_matriz.append(['', '', '', '', '', ''])
    
    # DESPESAS - RESUMIDAS POR CATEGORIA (SEM DETALHES INDIVIDUAIS)
    dados_matriz.append(['DESPESAS', '', '', '', '', ''])
    
    # Ordena por categoria
    categorias_ordenadas = ['ADS', 'Impostos', 'Juridico', 'Pro labore', 'Serviços', 'Sistemas']
    
    for categoria in categorias_ordenadas:
        if categoria in df_despesas['Categoria'].values:
            # Apenas a linha da categoria (sem itens individuais)
            linha_categoria = [f'{categoria}', '']
            total_categoria = 0
            
            for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
                valor = df_despesas[
                    (df_despesas['Data'].dt.strftime('%Y-%m') == mes) & 
                    (df_despesas['Categoria'] == categoria)
                ]['Valor'].sum()
                linha_categoria.append(formatar_moeda_br(valor) if valor > 0 else '-')
                total_categoria += valor
            
            linha_categoria.append(formatar_moeda_br(total_categoria))
            dados_matriz.append(linha_categoria)
    
    # Total despesas - usa valor da planilha original que já tem o total correto
    try:
        df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
        linha_total_original = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
        
        linha_total_desp = ['TOTAL DESPESAS', '']
        total_geral_desp = 0
        
        if not linha_total_original.empty:
            # Usa os valores corretos da planilha original
            for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
                # Encontra a coluna correspondente ao mês
                colunas_data = [col for col in df_matriz_original.columns if isinstance(col, pd.Timestamp)]
                coluna_mes = None
                for col in colunas_data:
                    if col.strftime('%Y-%m') == mes:
                        coluna_mes = col
                        break
                
                if coluna_mes is not None:
                    valor = linha_total_original[coluna_mes].iloc[0]
                    if pd.notna(valor) and valor != 0:
                        linha_total_desp.append(formatar_moeda_br(valor))
                        total_geral_desp += valor
                    else:
                        linha_total_desp.append(formatar_moeda_br(0))
                else:
                    linha_total_desp.append(formatar_moeda_br(0))
        else:
            # Fallback: calcula manualmente se não encontrar a linha
            for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
                df_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                df_principais = df_mes[~df_mes['Categoria'].str.startswith('  ', na=False)]
                valor = df_principais['Valor'].sum()
                linha_total_desp.append(formatar_moeda_br(valor))
                total_geral_desp += valor
        
        linha_total_desp.append(formatar_moeda_br(total_geral_desp))
    except Exception as e:
        st.error(f"Erro ao carregar totais da matriz: {e}")
        # Fallback em caso de erro
        linha_total_desp = ['TOTAL DESPESAS', ''] + [formatar_moeda_br(0)] * 5
    dados_matriz.append(linha_total_desp)
    
    # Separador
    dados_matriz.append(['', '', '', '', '', ''])
    
    # RESULTADO LÍQUIDO (Receitas - Custos - Despesas)
    linha_resultado = ['RESULTADO LÍQUIDO', '']
    resultado_total = 0
    
    for i, mes in enumerate(['2025-06', '2025-07', '2025-08', '2025-09']):
        rec = df_receitas[df_receitas['Data'].dt.strftime('%Y-%m') == mes]['Valor'].sum()
        
        # Calcula TODOS os custos (adquirente + antecipação)
        custo_adquirente = receitas_por_mes.get(mes, 0) * 0.025  # Todos os meses
        custo_antecipacao = tarifas_antecipacao.get(mes, 0)      # Só setembro em diante
        custos_total = custo_adquirente + custo_antecipacao
        
        # Usa valor do TOTAL DESPESAS da planilha original
        try:
            df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
            linha_total_original = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
            
            desp = 0
            if not linha_total_original.empty:
                # Encontra a coluna correspondente ao mês
                colunas_data = [col for col in df_matriz_original.columns if isinstance(col, pd.Timestamp)]
                for col in colunas_data:
                    if col.strftime('%Y-%m') == mes:
                        valor_total = linha_total_original[col].iloc[0]
                        if pd.notna(valor_total):
                            desp = valor_total
                        break
            
            if desp == 0:
                # Fallback: calcula manualmente
                df_desp_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                df_desp_principais = df_desp_mes[~df_desp_mes['Categoria'].str.startswith('  ', na=False)]
                desp = df_desp_principais['Valor'].sum()
        except:
            # Fallback em caso de erro
            df_desp_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
            df_desp_principais = df_desp_mes[~df_desp_mes['Categoria'].str.startswith('  ', na=False)]
            desp = df_desp_principais['Valor'].sum()
        
        resultado = rec - custos_total - desp
        linha_resultado.append(formatar_moeda_br(resultado))
        resultado_total += resultado
    
    linha_resultado.append(formatar_moeda_br(resultado_total))
    dados_matriz.append(linha_resultado)
    
    # Cria DataFrame da matriz
    df_matriz = pd.DataFrame(dados_matriz, columns=[
        'Categoria', 'Status', 'Jun/25', 'Jul/25', 'Ago/25', 'Set/25', 'Total'
    ])
    
    # Exibe a matriz
    st.dataframe(
        df_matriz,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Categoria": st.column_config.TextColumn("Categoria", width=300),
            "Status": st.column_config.TextColumn("Status", width=100),
            "Jun/25": st.column_config.TextColumn("Jun/25", width=120),
            "Jul/25": st.column_config.TextColumn("Jul/25", width=120),
            "Ago/25": st.column_config.TextColumn("Ago/25", width=120),
            "Set/25": st.column_config.TextColumn("Set/25", width=120),
            "Total": st.column_config.TextColumn("Total", width=140)
        }
    )
    
    st.info("**Visão resumida**: Despesas agrupadas por categoria. Para ver itens individuais, acesse 'Matriz Realizada'.")

def mostrar_matriz_realizada(df_despesas, df_receitas):
    """Matriz detalhada com exibição completa das despesas individuais"""
    
    st.header("📊 Realizado")
    
    # Cria dados consolidados
    dados_matriz = []
    
    # Header
    meses = ['Jun/25', 'Jul/25', 'Ago/25', 'Set/25*']
    
    # RECEITAS
    
    # Receitas por status
    for status in ['Realizado', 'Projetado']:
        linha = [f'Receitas {status}', '']
        total_status = 0
        
        for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
            valor = df_receitas[
                (df_receitas['Data'].dt.strftime('%Y-%m') == mes) & 
                (df_receitas['Status'] == status)
            ]['Valor'].sum()
            linha.append(formatar_moeda_br(valor))
            total_status += valor
        
        linha.append(formatar_moeda_br(total_status))
        dados_matriz.append(linha)
    
    # Total receitas
    linha_total_rec = ['TOTAL RECEITAS', '']
    total_geral_rec = 0
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        valor = df_receitas[df_receitas['Data'].dt.strftime('%Y-%m') == mes]['Valor'].sum()
        linha_total_rec.append(formatar_moeda_br(valor))
        total_geral_rec += valor
    linha_total_rec.append(formatar_moeda_br(total_geral_rec))
    dados_matriz.append(linha_total_rec)
    
    # Separador
    dados_matriz.append(['', '', '', '', '', ''])
    
    # CUSTOS (Tarifas)
    dados_matriz.append(['CUSTOS', '', '', '', '', ''])
    
    # Receitas brutas por mês para cálculo das tarifas
    receitas_por_mes = {
        '2025-06': 198198.40,  # Junho bruto
        '2025-07': 433176.64,  # Julho bruto  
        '2025-08': 146682.25,  # Agosto bruto
        '2025-09': 22278.47    # Setembro bruto (realizado)
    }
    
    # Tarifa de adquirente (2.5% sobre receitas - TODOS os meses)
    linha_adquirente = ['Tarifa Adquirente (2.5%)', '']
    total_adquirente = 0
    
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        if mes in receitas_por_mes:
            tarifa = receitas_por_mes[mes] * 0.025
            linha_adquirente.append(formatar_moeda_br(tarifa))
            total_adquirente += tarifa
        else:
            linha_adquirente.append('-')
    linha_adquirente.append(formatar_moeda_br(total_adquirente))
    dados_matriz.append(linha_adquirente)
    
    # Tarifa de antecipação (3.33% apenas setembro em diante)
    linha_antecipacao = ['Tarifa Antecipação (3.33%)', '']
    total_antecipacao = 0
    
    tarifas_antecipacao = {
        '2025-06': 0,           # Não teve
        '2025-07': 0,           # Não teve
        '2025-08': 0,           # Não teve
        '2025-09': receitas_por_mes['2025-09'] * 0.0333  # Setembro: 3.33% sobre realizado
    }
    
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        tarifa = tarifas_antecipacao.get(mes, 0)
        if tarifa > 0:
            linha_antecipacao.append(formatar_moeda_br(tarifa))
            total_antecipacao += tarifa
        else:
            linha_antecipacao.append('-')
    linha_antecipacao.append(formatar_moeda_br(total_antecipacao))
    dados_matriz.append(linha_antecipacao)
    
    # Total custos (adquirente + antecipação)
    linha_total_custos = ['TOTAL CUSTOS', '']
    total_custos_geral = 0
    
    for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
        # Adquirente todos os meses
        custo_adquirente = receitas_por_mes.get(mes, 0) * 0.025
        # Antecipação só setembro em diante
        custo_antecipacao = tarifas_antecipacao.get(mes, 0)
        
        total_mes = custo_adquirente + custo_antecipacao
        linha_total_custos.append(formatar_moeda_br(total_mes))
        total_custos_geral += total_mes
    
    linha_total_custos.append(formatar_moeda_br(total_custos_geral))
    dados_matriz.append(linha_total_custos)
    
    # Separador
    dados_matriz.append(['', '', '', '', '', ''])
    
    # DESPESAS - Exibição completa como na exportação
    dados_matriz.append(['DESPESAS', '', '', '', '', ''])
    
    # Ordena por categoria e depois por descrição para exibição organizada
    categorias_ordenadas = ['ADS', 'Impostos', 'Juridico', 'Pro labore', 'Serviços', 'Sistemas']
    
    for categoria in categorias_ordenadas:
        if categoria in df_despesas['Categoria'].values:
            # Header da categoria
            dados_matriz.append([f'📂 {categoria.upper()}', '', '', '', '', ''])
            
            # Itens individuais da categoria (apenas únicos)
            descricoes_unicas = df_despesas[df_despesas['Categoria'] == categoria]['Descricao'].unique()
            descricoes_unicas = sorted(descricoes_unicas)
            
            for descricao in descricoes_unicas:
                # Pega status do primeiro registro dessa descrição
                item_info = df_despesas[df_despesas['Descricao'] == descricao].iloc[0]
                linha = [f'  → {descricao}', item_info['Status']]
                total_item = 0
                
                for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
                    valor = df_despesas[
                        (df_despesas['Data'].dt.strftime('%Y-%m') == mes) & 
                        (df_despesas['Descricao'] == descricao)
                    ]['Valor'].sum()
                    linha.append(formatar_moeda_br(valor) if valor > 0 else '-')
                    total_item += valor
                
                linha.append(formatar_moeda_br(total_item) if total_item > 0 else '-')
                dados_matriz.append(linha)
            
            # Subtotal da categoria
            linha_subtotal = [f'SUBTOTAL {categoria}', '']
            total_categoria = 0
            
            for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
                valor = df_despesas[
                    (df_despesas['Data'].dt.strftime('%Y-%m') == mes) & 
                    (df_despesas['Categoria'] == categoria)
                ]['Valor'].sum()
                linha_subtotal.append(formatar_moeda_br(valor) if valor > 0 else '-')
                total_categoria += valor
            
            linha_subtotal.append(formatar_moeda_br(total_categoria))
            dados_matriz.append(linha_subtotal)
            
            # Separador entre categorias
            dados_matriz.append(['', '', '', '', '', ''])
    
    # Total despesas - usa valor da planilha original que já tem o total correto
    try:
        df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
        linha_total_original = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
        
        linha_total_desp = ['TOTAL DESPESAS', '']
        total_geral_desp = 0
        
        if not linha_total_original.empty:
            # Usa os valores corretos da planilha original
            for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
                # Encontra a coluna correspondente ao mês
                colunas_data = [col for col in df_matriz_original.columns if isinstance(col, pd.Timestamp)]
                coluna_mes = None
                for col in colunas_data:
                    if col.strftime('%Y-%m') == mes:
                        coluna_mes = col
                        break
                
                if coluna_mes is not None:
                    valor = linha_total_original[coluna_mes].iloc[0]
                    if pd.notna(valor) and valor != 0:
                        linha_total_desp.append(formatar_moeda_br(valor))
                        total_geral_desp += valor
                    else:
                        linha_total_desp.append(formatar_moeda_br(0))
                else:
                    linha_total_desp.append(formatar_moeda_br(0))
        else:
            # Fallback: calcula manualmente se não encontrar a linha
            for mes in ['2025-06', '2025-07', '2025-08', '2025-09']:
                df_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                df_principais = df_mes[~df_mes['Categoria'].str.startswith('  ', na=False)]
                valor = df_principais['Valor'].sum()
                linha_total_desp.append(formatar_moeda_br(valor))
                total_geral_desp += valor
        
        linha_total_desp.append(formatar_moeda_br(total_geral_desp))
    except Exception as e:
        st.error(f"Erro ao carregar totais da matriz: {e}")
        # Fallback em caso de erro
        linha_total_desp = ['TOTAL DESPESAS', ''] + [formatar_moeda_br(0)] * 5
    dados_matriz.append(linha_total_desp)
    
    # Separador
    dados_matriz.append(['', '', '', '', '', ''])
    
    # RESULTADO LÍQUIDO (Receitas - Custos - Despesas)
    linha_resultado = ['RESULTADO LÍQUIDO', '']
    resultado_total = 0
    
    for i, mes in enumerate(['2025-06', '2025-07', '2025-08', '2025-09']):
        rec = df_receitas[df_receitas['Data'].dt.strftime('%Y-%m') == mes]['Valor'].sum()
        
        # Calcula TODOS os custos (adquirente + antecipação)
        custo_adquirente = receitas_por_mes.get(mes, 0) * 0.025  # Todos os meses
        custo_antecipacao = tarifas_antecipacao.get(mes, 0)      # Só setembro em diante
        custos_total = custo_adquirente + custo_antecipacao
        
        # Usa valor do TOTAL DESPESAS da planilha original
        try:
            df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
            linha_total_original = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
            
            desp = 0
            if not linha_total_original.empty:
                # Encontra a coluna correspondente ao mês
                colunas_data = [col for col in df_matriz_original.columns if isinstance(col, pd.Timestamp)]
                for col in colunas_data:
                    if col.strftime('%Y-%m') == mes:
                        valor_total = linha_total_original[col].iloc[0]
                        if pd.notna(valor_total):
                            desp = valor_total
                        break
            
            if desp == 0:
                # Fallback: calcula manualmente
                df_desp_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                df_desp_principais = df_desp_mes[~df_desp_mes['Categoria'].str.startswith('  ', na=False)]
                desp = df_desp_principais['Valor'].sum()
        except:
            # Fallback em caso de erro
            df_desp_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
            df_desp_principais = df_desp_mes[~df_desp_mes['Categoria'].str.startswith('  ', na=False)]
            desp = df_desp_principais['Valor'].sum()
        
        resultado = rec - custos_total - desp
        linha_resultado.append(formatar_moeda_br(resultado))
        resultado_total += resultado
    
    linha_resultado.append(formatar_moeda_br(resultado_total))
    dados_matriz.append(linha_resultado)
    
    # Cria DataFrame da matriz
    colunas = ['Categoria', ''] + meses + ['Total']
    df_matriz = pd.DataFrame(dados_matriz, columns=colunas)
    
    # Aplica formatação condicional
    def estilizar_tabela(row):
        estilos = [''] * len(row)
        
        # Headers principais em negrito
        if 'TOTAL' in str(row.iloc[0]) or 'RESULTADO' in str(row.iloc[0]):
            estilos = ['font-weight: bold; background-color: rgba(128, 128, 128, 0.2)'] * len(row)
        
        # Valores negativos em vermelho
        for i in range(2, len(row)):
            valor_str = str(row.iloc[i])
            if valor_str.startswith('R$') and '-' in valor_str:
                estilos[i] = 'color: #dc3545'
        
        return estilos
    
    # Configuração das colunas para melhor alinhamento
    col_config = {
        'Categoria': st.column_config.TextColumn('Categoria', width='large'),
        '': st.column_config.TextColumn('', width='small'),
    }
    
    # Colunas de meses e total com largura média
    for mes in meses + ['Total']:
        col_config[mes] = st.column_config.TextColumn(
            mes,
            width='medium',
            help=f'Valores de {mes}'
        )
    
    # Exibe com altura maior para melhor visualização
    st.dataframe(
        df_matriz.style.apply(estilizar_tabela, axis=1),
        use_container_width=True,
        hide_index=True,
        height=600,
        column_config=col_config
    )
    
    

def exportar_matriz_excel(df_despesas, df_receitas, tarifas_antecipacao):
    """Exporta matriz financeira detalhada para Excel"""
    
    try:
        # Criar arquivo Excel em memória
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Preparar dados da matriz detalhada
            dados_detalhados = []
            
            meses = ['Jun/25', 'Jul/25', 'Ago/25', 'Set/25']
            meses_str = ['2025-06', '2025-07', '2025-08', '2025-09']
            
            # RECEITAS
            dados_detalhados.append(['RECEITAS', '', '', '', '', ''])
            
            # Receitas por status
            for status in ['Realizado', 'Projetado']:
                linha = [f'Receitas {status}', '']
                total_status = 0
                
                for mes in meses_str:
                    valor = df_receitas[
                        (df_receitas['Data'].dt.strftime('%Y-%m') == mes) & 
                        (df_receitas['Status'] == status)
                    ]['Valor'].sum()
                    linha.append(valor)
                    total_status += valor
                
                linha.append(total_status)
                dados_detalhados.append(linha)
            
            # Total receitas
            linha_total_rec = ['TOTAL RECEITAS', '']
            for mes in meses_str:
                valor = df_receitas[df_receitas['Data'].dt.strftime('%Y-%m') == mes]['Valor'].sum()
                linha_total_rec.append(valor)
            linha_total_rec.append(sum(linha_total_rec[2:]))
            dados_detalhados.append(linha_total_rec)
            
            dados_detalhados.append(['', '', '', '', '', ''])  # Separador
            
            # CUSTOS
            dados_detalhados.append(['CUSTOS', '', '', '', '', ''])
            
            # Receitas brutas para cálculo
            receitas_por_mes_export = {
                '2025-06': 198198.40,  # Junho bruto
                '2025-07': 433176.64,  # Julho bruto  
                '2025-08': 146682.25,  # Agosto bruto
                '2025-09': 22278.47    # Setembro bruto (realizado)
            }
            
            # Tarifa adquirente (todos os meses)
            linha_adquirente = ['Tarifa Adquirente (2.5%)', '']
            for mes in meses_str:
                tarifa = receitas_por_mes_export.get(mes, 0) * 0.025
                linha_adquirente.append(tarifa)
            linha_adquirente.append(sum(linha_adquirente[2:]))
            dados_detalhados.append(linha_adquirente)
            
            # Tarifa antecipação (só setembro em diante)
            tarifas_antecipacao_export = {
                '2025-06': 0,
                '2025-07': 0,
                '2025-08': 0,
                '2025-09': receitas_por_mes_export['2025-09'] * 0.0333
            }
            
            linha_antecip = ['Tarifa Antecipação (3.33%)', '']
            for mes in meses_str:
                tarifa = tarifas_antecipacao_export.get(mes, 0)
                linha_antecip.append(tarifa)
            linha_antecip.append(sum(linha_antecip[2:]))
            dados_detalhados.append(linha_antecip)
            
            # Total custos
            linha_total_custos_export = ['TOTAL CUSTOS', '']
            for i, mes in enumerate(meses_str):
                custo_adq = linha_adquirente[i+2]
                custo_ant = linha_antecip[i+2]
                linha_total_custos_export.append(custo_adq + custo_ant)
            linha_total_custos_export.append(sum(linha_total_custos_export[2:]))
            dados_detalhados.append(linha_total_custos_export)
            
            dados_detalhados.append(['', '', '', '', '', ''])  # Separador
            
            # DESPESAS DETALHADAS
            dados_detalhados.append(['DESPESAS', '', '', '', '', ''])
            
            # Por categoria
            categorias = sorted(df_despesas['Categoria'].unique())
            for categoria in categorias:
                # Linha da categoria
                linha_cat = [f'{categoria}', '']
                total_cat = 0
                valores_mes = []
                
                for mes in meses_str:
                    valor = df_despesas[
                        (df_despesas['Data'].dt.strftime('%Y-%m') == mes) & 
                        (df_despesas['Categoria'] == categoria)
                    ]['Valor'].sum()
                    linha_cat.append(valor)
                    valores_mes.append(valor)
                    total_cat += valor
                
                linha_cat.append(total_cat)
                dados_detalhados.append(linha_cat)
                
                # Detalhes dos itens da categoria
                itens = df_despesas[df_despesas['Categoria'] == categoria]['Descricao'].unique()
                for item in itens:
                    linha_item = [f'  → {item}', '']
                    for mes in meses_str:
                        valor = df_despesas[
                            (df_despesas['Data'].dt.strftime('%Y-%m') == mes) & 
                            (df_despesas['Categoria'] == categoria) &
                            (df_despesas['Descricao'] == item)
                        ]['Valor'].sum()
                        linha_item.append(valor if valor > 0 else 0)
                    linha_item.append(sum(linha_item[2:]))
                    dados_detalhados.append(linha_item)
            
            # Total despesas - usa valor da planilha original
            linha_total_desp = ['TOTAL DESPESAS', '']
            try:
                df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
                linha_total_original = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
                
                if not linha_total_original.empty:
                    # Usa os valores corretos da planilha original
                    for mes in meses_str:
                        # Encontra a coluna correspondente ao mês
                        colunas_data = [col for col in df_matriz_original.columns if isinstance(col, pd.Timestamp)]
                        valor = 0
                        for col in colunas_data:
                            if col.strftime('%Y-%m') == mes:
                                valor_total = linha_total_original[col].iloc[0]
                                if pd.notna(valor_total):
                                    valor = valor_total
                                break
                        linha_total_desp.append(valor)
                else:
                    # Fallback: calcula manualmente
                    for mes in meses_str:
                        df_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                        df_principais = df_mes[~df_mes['Categoria'].str.startswith('  ', na=False)]
                        valor = df_principais['Valor'].sum()
                        linha_total_desp.append(valor)
            except Exception as e:
                # Fallback em caso de erro
                for mes in meses_str:
                    df_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                    df_principais = df_mes[~df_mes['Categoria'].str.startswith('  ', na=False)]
                    valor = df_principais['Valor'].sum()
                    linha_total_desp.append(valor)
            
            linha_total_desp.append(sum(linha_total_desp[2:]))
            dados_detalhados.append(linha_total_desp)
            
            dados_detalhados.append(['', '', '', '', '', ''])  # Separador
            
            # RESULTADO LÍQUIDO
            linha_resultado = ['RESULTADO LÍQUIDO', '']
            for i, mes in enumerate(meses_str):
                rec = df_receitas[df_receitas['Data'].dt.strftime('%Y-%m') == mes]['Valor'].sum()
                # Custos totais (adquirente + antecipação)
                custos_total = linha_total_custos_export[i+2]
                # Usa valor do TOTAL DESPESAS da planilha original
                try:
                    df_matriz_original = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
                    linha_total_original = df_matriz_original[df_matriz_original['Categoria'] == 'TOTAL DESPESAS']
                    
                    desp = 0
                    if not linha_total_original.empty:
                        # Encontra a coluna correspondente ao mês
                        colunas_data = [col for col in df_matriz_original.columns if isinstance(col, pd.Timestamp)]
                        for col in colunas_data:
                            if col.strftime('%Y-%m') == mes:
                                valor_total = linha_total_original[col].iloc[0]
                                if pd.notna(valor_total):
                                    desp = valor_total
                                break
                    
                    if desp == 0:
                        # Fallback: calcula manualmente
                        df_desp_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                        df_desp_principais = df_desp_mes[~df_desp_mes['Categoria'].str.startswith('  ', na=False)]
                        desp = df_desp_principais['Valor'].sum()
                except:
                    # Fallback em caso de erro
                    df_desp_mes = df_despesas[df_despesas['Data'].dt.strftime('%Y-%m') == mes]
                    df_desp_principais = df_desp_mes[~df_desp_mes['Categoria'].str.startswith('  ', na=False)]
                    desp = df_desp_principais['Valor'].sum()
                resultado = rec - custos_total - desp
                linha_resultado.append(resultado)
            linha_resultado.append(sum(linha_resultado[2:]))
            dados_detalhados.append(linha_resultado)
            
            # Criar DataFrame
            colunas = ['Categoria', 'Tipo'] + meses + ['Total']
            df_export = pd.DataFrame(dados_detalhados, columns=colunas)
            
            # Exportar para Excel
            df_export.to_excel(writer, sheet_name='Matriz Detalhada', index=False)
            
            # Formatar o Excel
            workbook = writer.book
            worksheet = writer.sheets['Matriz Detalhada']
            
            # Formatar cabeçalho
            for col in range(1, 8):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            
            # Formatar linhas de totais
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=1)
                if cell.value and ('TOTAL' in str(cell.value) or 'RESULTADO' in str(cell.value)):
                    for col in range(1, 8):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='f0f0f0', end_color='f0f0f0', fill_type='solid')
            
            # Ajustar largura das colunas
            worksheet.column_dimensions['A'].width = 35
            for col in ['B', 'C', 'D', 'E', 'F', 'G']:
                worksheet.column_dimensions[col].width = 15
        
        # Preparar download
        output.seek(0)
        st.download_button(
            label="💾 Baixar Arquivo Excel",
            data=output,
            file_name=f"matriz_financeira_detalhada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.success("✅ Arquivo pronto para download! Clique no botão acima.")
        
    except Exception as e:
        st.error(f"❌ Erro ao gerar arquivo: {e}")

def mostrar_fluxo_caixa_projecoes():
    """Análise Completa de Fluxo de Caixa com filtros e visualizações"""
    st.header("Fluxo de Caixa Detalhado")
    
    # Criar abas para diferentes análises
    tab1, tab2, tab3 = st.tabs(["Resumo", "Análise Temporal", "Projeções"])
    
    with tab1:
        mostrar_resumo_fluxo_caixa()
    
    with tab2:
        mostrar_analise_temporal_fluxo()
    
    
    with tab3:
        mostrar_projecoes_originais_fluxo()

def mostrar_resumo_fluxo_caixa():
    """Mostra resumo consolidado do fluxo de caixa com dados reais"""
    st.subheader("Resumo Consolidado")

    try:
        # Carrega dados reais
        saldos_reais = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Saldos Adquirentes')
        resumo_real = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Resumo')

        # Dados do fluxo atual
        saldo_atual = saldos_reais['Saldo Disponível'].sum()
        a_receber = saldos_reais['A Receber'].sum()

        # Métricas principais
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Saldo Atual", formatar_valor_compacto(saldo_atual))

        with col2:
            st.metric("A Receber", formatar_valor_compacto(a_receber))

        with col3:
            saldo_projetado = saldo_atual + a_receber
            st.metric("Saldo Projetado", formatar_valor_compacto(saldo_projetado))

        with col4:
            st.metric("Dias de Caixa", "~90 dias")

        # Distribuição por adquirente
        st.subheader("Distribuição por Adquirente")

        col1, col2 = st.columns(2)

        with col1:
            st.info("Gráfico em manutenção")

        with col2:
            # Tabela detalhada formatada
            saldos_formatados = saldos_reais.copy()
            for col in ["Saldo Disponível", "A Receber", "Total"]:
                if col in saldos_formatados.columns:
                    saldos_formatados[col] = saldos_formatados[col].apply(lambda x: f"R$ {x:,.2f}")
            st.dataframe(saldos_formatados, width='stretch')

    except Exception as e:
        st.error(f"Erro ao carregar resumo: {e}")
def mostrar_analise_temporal_fluxo():
    """Análise temporal com filtros de período"""
    st.subheader("Análise Temporal do Fluxo")
    
    # Filtros de período
    col1, col2, col3 = st.columns(3)
    
    with col1:
        periodo = st.selectbox(
            "Período de Análise",
            ["7 dias", "30 dias", "60 dias", "90 dias", "180 dias"],
            index=1
        )
    
    with col2:
        data_inicio = st.date_input("Data Inicial", value=pd.Timestamp.now().date())
    
    with col3:
        incluir_projecoes = st.checkbox("Incluir Projeções", value=True)
    
    # Processar dados baseado no filtro
    dias = int(periodo.split()[0])
    data_final = pd.Timestamp(data_inicio) + pd.Timedelta(days=dias)
    
    # Gerar dados do fluxo para o período
    fluxo_temporal = gerar_fluxo_temporal(data_inicio, data_final, incluir_projecoes)
    
    if not fluxo_temporal.empty:
        # Gráfico de linha do saldo acumulado
        fig_linha = px.line(
            fluxo_temporal,
            x='Data',
            y='Saldo_Acumulado',
            title=f"Evolução do Fluxo de Caixa - {periodo}",
            markers=True
        )
        fig_linha.update_layout(
            xaxis_title="Data",
            yaxis_title="Saldo (R$)",
            hovermode='x unified'
        )
        st.plotly_chart(fig_linha, use_container_width=True)
        
        # Métricas do período
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Saldo Inicial", formatar_valor_compacto(fluxo_temporal['Saldo_Acumulado'].iloc[0]))
        
        with col2:
            st.metric("Saldo Final", formatar_valor_compacto(fluxo_temporal['Saldo_Acumulado'].iloc[-1]))
        
        with col3:
            variacao = fluxo_temporal['Saldo_Acumulado'].iloc[-1] - fluxo_temporal['Saldo_Acumulado'].iloc[0]
            st.metric("Variação", formatar_valor_compacto(variacao))
        
        with col4:
            menor_saldo = fluxo_temporal['Saldo_Acumulado'].min()
            st.metric("Menor Saldo", formatar_valor_compacto(menor_saldo))
    
    else:
        st.warning("Nenhum dado encontrado para o período selecionado")

def mostrar_graficos_fluxo_caixa():
    """Gráficos detalhados do fluxo de caixa"""
    st.subheader("Visualizações do Fluxo de Caixa")
    
    try:
        # Gera fluxo temporal completo para visualizações
        data_inicio = pd.Timestamp.now().date()
        data_final = data_inicio + pd.Timedelta(days=90)
        fluxo_completo = gerar_fluxo_temporal(data_inicio, data_final, incluir_projecoes=True)
        
        if not fluxo_completo.empty:
            # Gráfico de barras das entradas e saídas
            fig_bar = go.Figure()
            
            # Adicionar entradas
            entradas = fluxo_completo[fluxo_completo['Fluxo_Liquido'] > 0]
            if not entradas.empty:
                fig_bar.add_trace(go.Bar(
                    x=entradas['Data'],
                    y=entradas['Fluxo_Liquido'],
                    name='Entradas',
                    marker_color='green'
                ))
            
            # Adicionar saídas
            saidas = fluxo_completo[fluxo_completo['Fluxo_Liquido'] < 0]
            if not saidas.empty:
                fig_bar.add_trace(go.Bar(
                    x=saidas['Data'],
                    y=saidas['Fluxo_Liquido'],
                    name='Saídas',
                    marker_color='red'
                ))
            
            fig_bar.update_layout(
                title="Entradas e Saídas Diárias",
                xaxis_title="Data",
                yaxis_title="Valor (R$)",
                barmode='relative'
            )
            st.plotly_chart(fig_bar, use_container_width=True)
            
            # Gráfico de área do saldo acumulado
            fig_area = px.area(
                fluxo_completo,
                x='Data',
                y='Saldo_Acumulado',
                title="Evolução do Saldo Acumulado",
                color_discrete_sequence=['lightblue']
            )
            st.plotly_chart(fig_area, use_container_width=True)
        
    except Exception as e:
        st.error(f"Erro ao gerar gráficos: {e}")

def mostrar_dados_detalhados_fluxo():
    """Tabela detalhada do fluxo de caixa"""
    st.subheader("Dados Detalhados")
    
    # Filtros
    col1, col2 = st.columns(2)
    
    with col1:
        periodo = st.selectbox(
            "Período",
            ["30 dias", "60 dias", "90 dias", "180 dias"],
            key="detalhes_periodo"
        )
    
    with col2:
        mostrar_apenas = st.selectbox(
            "Mostrar",
            ["Todos", "Apenas Entradas", "Apenas Saídas", "Apenas Recebimentos"],
            key="detalhes_filtro"
        )
    
    # Gerar dados
    dias = int(periodo.split()[0])
    data_inicio = pd.Timestamp.now().date()
    data_final = data_inicio + pd.Timedelta(days=dias)
    
    fluxo_detalhado = gerar_fluxo_temporal(data_inicio, data_final, incluir_projecoes=True)
    
    if not fluxo_detalhado.empty:
        # Aplicar filtros
        if mostrar_apenas == "Apenas Entradas":
            fluxo_detalhado = fluxo_detalhado[fluxo_detalhado['Fluxo_Liquido'] > 0]
        elif mostrar_apenas == "Apenas Saídas":
            fluxo_detalhado = fluxo_detalhado[fluxo_detalhado['Fluxo_Liquido'] < 0]
        elif mostrar_apenas == "Apenas Recebimentos":
            fluxo_detalhado = fluxo_detalhado[fluxo_detalhado['Tipo'] == 'Recebimento']
        
        # Adicionar colunas formatadas
        fluxo_detalhado['Valor_Formatado'] = fluxo_detalhado['Fluxo_Liquido'].apply(formatar_moeda_br)
        fluxo_detalhado['Saldo_Formatado'] = fluxo_detalhado['Saldo_Acumulado'].apply(formatar_moeda_br)
        
        # Exibir tabela
        colunas_exibir = ['Data', 'Descricao', 'Valor_Formatado', 'Saldo_Formatado', 'Tipo']
        st.dataframe(
            fluxo_detalhado[colunas_exibir],
            use_container_width=True,
            column_config={
                'Data': st.column_config.DateColumn('Data'),
                'Descricao': 'Descrição',
                'Valor_Formatado': 'Valor',
                'Saldo_Formatado': 'Saldo Acumulado',
                'Tipo': 'Tipo'
            }
        )
        
        # Resumo da consulta
        st.subheader("Resumo do Período")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_entradas = fluxo_detalhado[fluxo_detalhado['Fluxo_Liquido'] > 0]['Fluxo_Liquido'].sum()
            st.metric("Total Entradas", formatar_valor_compacto(total_entradas))
        
        with col2:
            total_saidas = fluxo_detalhado[fluxo_detalhado['Fluxo_Liquido'] < 0]['Fluxo_Liquido'].sum()
            st.metric("Total Saídas", formatar_valor_compacto(abs(total_saidas)))
        
        with col3:
            saldo_liquido = total_entradas + total_saidas
            st.metric("Saldo Líquido", formatar_valor_compacto(saldo_liquido))
    
    else:
        st.warning("Nenhum dado encontrado para o período selecionado")

def gerar_fluxo_temporal(data_inicio, data_final, incluir_projecoes=True):
    """Gera dados do fluxo de caixa combinando dados reais + projeções da matriz"""
    try:
        # 1. Carrega dados reais atuais
        saldos_reais = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Saldos Adquirentes')
        resumo_real = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Resumo')
        
        saldo_disponivel = saldos_reais['Saldo Disponível'].sum()
        saldo_a_receber = saldos_reais['A Receber'].sum()
        saldo_inicial = saldo_disponivel + saldo_a_receber
        a_receber = saldos_reais['A Receber'].sum()
        
        # 2. Carrega dados das projeções da matriz (se incluir_projecoes=True)
        dados_matriz = {}
        if incluir_projecoes:
            try:
                df_matriz = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
                
                # Identifica colunas de data na matriz
                colunas_data = []
                for col in df_matriz.columns:
                    if isinstance(col, (pd.Timestamp, datetime)):
                        colunas_data.append(col)
                
                # Processa dados mensais da matriz
                for col in colunas_data:
                    mes_ano = col.strftime('%Y-%m')
                    
                    # Extrai receitas, custos e despesas da matriz
                    receitas = 0
                    custos = 0
                    despesas = 0
                    
                    # RECEITAS
                    receitas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL RECEITAS']
                    if not receitas_linha.empty:
                        receitas = converter_para_float_seguro(receitas_linha[col].iloc[0])
                    
                    # CUSTOS
                    custos_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL CUSTOS']
                    if not custos_linha.empty:
                        custos = converter_para_float_seguro(custos_linha[col].iloc[0])
                    
                    # DESPESAS
                    despesas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL DESPESAS']
                    if not despesas_linha.empty:
                        despesas = converter_para_float_seguro(despesas_linha[col].iloc[0])
                    
                    dados_matriz[mes_ano] = {
                        'receitas': receitas,
                        'custos': custos,
                        'despesas': despesas,
                        'fluxo_liquido': receitas - custos - despesas
                    }
                    
            except Exception as e:
                st.warning(f"Não foi possível carregar dados da matriz: {e}")
        
        # 3. Gerar fluxo diário combinado
        fluxo_diario = []
        data_atual = pd.Timestamp(data_inicio)
        data_fim = pd.Timestamp(data_final)
        saldo_acumulado = saldo_inicial
        
        while data_atual <= data_fim:
            valor_dia = 0
            descricao = "Sem movimentação"
            tipo = "Neutro"
            
            # Recebimentos reais do Asaas (primeiros 90 dias)
            dias_passados = (data_atual - pd.Timestamp(data_inicio)).days
            if dias_passados < 90 and a_receber > 0:
                valor_diario_asaas = a_receber / 90
                if data_atual.day % 3 == 0:  # A cada 3 dias
                    valor_dia += valor_diario_asaas * 3
                    descricao = "Recebimento Asaas (Real)"
                    tipo = "Recebimento Real"
            
            # Dados da matriz de projeções (distribuídos por dia no mês)
            mes_ano_atual = data_atual.strftime('%Y-%m')
            if mes_ano_atual in dados_matriz:
                dados_mes = dados_matriz[mes_ano_atual]
                dias_no_mes = pd.Timestamp(data_atual.year, data_atual.month, 1).days_in_month
                
                # Distribui receitas da matriz ao longo do mês
                if dados_mes['receitas'] > 0:
                    valor_dia += dados_mes['receitas'] / dias_no_mes
                    if descricao == "Sem movimentação":
                        descricao = "Receitas Projetadas"
                        tipo = "Receita Projetada"
                
                # Distribui custos e despesas da matriz
                if dados_mes['custos'] > 0 or dados_mes['despesas'] > 0:
                    valor_dia -= (dados_mes['custos'] + dados_mes['despesas']) / dias_no_mes
                    if tipo == "Neutro":
                        descricao = "Custos/Despesas Projetados"
                        tipo = "Despesa Projetada"
            
            saldo_acumulado += valor_dia
            
            fluxo_diario.append({
                'Data': data_atual.date(),
                'Descricao': descricao,
                'Fluxo_Liquido': valor_dia,
                'Saldo_Acumulado': saldo_acumulado,
                'Tipo': tipo
            })
            
            data_atual += pd.Timedelta(days=1)
        
        return pd.DataFrame(fluxo_diario)
        
    except Exception as e:
        st.error(f"Erro ao gerar fluxo temporal: {e}")
        return pd.DataFrame()

def mostrar_projecoes_originais_fluxo():
    """Mostra as projeções originais de 5 anos como estavam antes"""
    st.subheader("Projeções de 5 Anos - Dados Originais da Matriz")
    
    try:
        # Carrega dados DETALHADOS da matriz financeira (mês a mês)
        df_matriz = pd.read_excel('Matriz financeira.xlsx', sheet_name='Matriz Detalhada')
        
        # Identifica colunas de data
        colunas_data = []
        for col in df_matriz.columns:
            if isinstance(col, (pd.Timestamp, datetime)):
                colunas_data.append(col)
        
        # Cria range de meses das projeções (setembro/2025 em diante)  
        data_inicio = pd.to_datetime('2025-09-01')
        colunas_futuras = [col for col in colunas_data if col >= data_inicio]
        
        if not colunas_futuras:
            st.warning("Nenhum dado de projeção encontrado para setembro/2025 em diante")
            return
        
        # Processa dados mensais DETALHADOS
        fluxo_caixa = []
        
        # Saldo inicial atualizado com dados reais das adquirentes
        try:
            saldos_reais = pd.read_excel('relatorio_fluxo_caixa_2025-09-09 (6).xlsx', sheet_name='Saldos Adquirentes')
            saldo_disponivel = saldos_reais['Saldo Disponível'].sum()
            saldo_a_receber = saldos_reais['A Receber'].sum()
            saldo_inicial = saldo_disponivel + saldo_a_receber
        except:
            saldo_inicial = 653141.31  # Saldo inicial padrão
        
        saldo_acumulado = saldo_inicial
        
        for col in colunas_futuras:
            periodo = col
            
            # RECEITAS - pega TOTAL RECEITAS direto da planilha
            receitas_total = 0
            receitas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL RECEITAS']
            if not receitas_linha.empty:
                receitas_total = converter_para_float_seguro(receitas_linha[col].iloc[0])
            
            # CUSTOS - pega TOTAL CUSTOS direto da planilha
            custos_total = 0
            custos_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL CUSTOS']
            if not custos_linha.empty:
                custos_total = converter_para_float_seguro(custos_linha[col].iloc[0])
            
            # DESPESAS - pega TOTAL DESPESAS direto da planilha
            despesas_total = 0
            despesas_linha = df_matriz[df_matriz['Categoria'] == 'TOTAL DESPESAS']
            if not despesas_linha.empty:
                despesas_total = converter_para_float_seguro(despesas_linha[col].iloc[0])
            
            # IMPOSTOS - calcula baseado nas receitas mensais (Lucro Presumido)
            impostos_total = 0
            if col >= pd.to_datetime('2025-10-01') and receitas_total > 0:
                receitas_servicos = receitas_total * 0.4
                receitas_livros = receitas_total * 0.6

                # USA A MESMA FUNÇÃO PARA GARANTIR CONSISTÊNCIA
                impostos_mes_dict = calcular_impostos_brasileiros(receitas_servicos, receitas_livros, col)
                impostos_total = sum(impostos_mes_dict.values())

            # Calcula fluxo líquido do mês
            fluxo_liquido = receitas_total - custos_total - despesas_total - impostos_total
            saldo_acumulado += fluxo_liquido
            
            fluxo_caixa.append({
                'Período': periodo.strftime('%m/%Y'),
                'Data': periodo,
                'Receitas': receitas_total,
                'Custos': custos_total,
                'Despesas': despesas_total,
                'Fluxo_Líquido': fluxo_liquido,
                'Saldo_Acumulado': saldo_acumulado
            })
        
        df_fluxo = pd.DataFrame(fluxo_caixa)
        
        if not df_fluxo.empty:
            # Formatação dos valores
            def formatar_valor_visual(valor):
                if valor >= 0:
                    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                else:
                    return f"-R$ {abs(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Métricas principais
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.info(f"Período: {df_fluxo['Período'].iloc[0]} a {df_fluxo['Período'].iloc[-1]}")
            with col2:
                st.info(f"Saldo Inicial Atualizado: {formatar_valor_visual(saldo_inicial)}")
            with col3:
                st.info(f"Saldo Final: {formatar_valor_visual(df_fluxo['Saldo_Acumulado'].iloc[-1])}")
            with col4:
                fluxo_medio = df_fluxo['Fluxo_Líquido'].mean()
                st.info(f"Fluxo Médio: {formatar_valor_visual(fluxo_medio)}")
            
            # Prepara DataFrame para exibição com formatação
            df_display = df_fluxo.copy()
            
            # Formatar valores monetários
            for col in ['Receitas', 'Custos', 'Despesas', 'Fluxo_Líquido', 'Saldo_Acumulado']:
                df_display[f'{col}_Formatado'] = df_display[col].apply(formatar_valor_visual)
            
            # Colunas para exibir
            colunas_exibir = ['Período', 'Receitas_Formatado', 'Custos_Formatado', 'Despesas_Formatado', 
                            'Fluxo_Líquido_Formatado', 'Saldo_Acumulado_Formatado']
            
            df_exibicao = df_display[colunas_exibir].copy()
            df_exibicao.columns = ['Período', 'Receitas', 'Custos', 'Despesas', 'Fluxo Líquido', 'Saldo Acumulado']
            
            # Tabela principal
            st.subheader("Fluxo de Caixa Mensal Projetado")
            st.dataframe(df_exibicao, use_container_width=True)
            
            # Gráfico da evolução
            st.subheader("Evolução do Saldo Acumulado")
            fig = px.line(
                df_fluxo, 
                x='Data', 
                y='Saldo_Acumulado',
                title="Projeção do Saldo Acumulado",
                markers=True
            )
            fig.update_layout(
                xaxis_title="Período",
                yaxis_title="Saldo Acumulado (R$)",
                hovermode='x unified'
            )
            st.plotly_chart(fig, use_container_width=True)
            
            # Gráfico de barras do fluxo líquido
            st.subheader("Fluxo Líquido Mensal")
            fig_bar = px.bar(
                df_fluxo,
                x='Período',
                y='Fluxo_Líquido',
                title="Fluxo Líquido por Período",
                color='Fluxo_Líquido',
                color_continuous_scale=['red', 'yellow', 'green']
            )
            st.plotly_chart(fig_bar, use_container_width=True)
            
    except Exception as e:
        st.error(f"Erro ao processar projeções originais: {e}")

def converter_para_float_seguro(valor):
    """Converte valor para float de forma segura"""
    if pd.isna(valor) or valor == '' or valor == '-':
        return 0.0
    try:
        if isinstance(valor, str):
            valor = valor.replace('R$', '').replace('.', '').replace(',', '.').strip()
        return float(valor)
    except:
        return 0.0


def mostrar_contas_receber_asaas():
    """Mostra a seção de contas a receber do Asaas com competência vs caixa"""
    st.header("Contas a Receber - Asaas")
    st.markdown("**Análise de Competência vs Caixa com reconciliação de vendas e recebimentos**")
    
    # Instanciar o módulo
    contas_asaas = ContasReceberAsaas()
    
    # Upload de arquivos
    st.subheader("Upload dos Arquivos")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Arquivo de Cobranças (Vendas)**")
        arquivo_cobrancas = st.file_uploader(
            "Upload arquivo de cobranças",
            type=['xlsx'], 
            key="cobrancas_asaas",
            help="Arquivo com as vendas/cobranças geradas no Asaas"
        )
        
    with col2:
        st.markdown("**Arquivo do Extrato Asaas**")  
        arquivo_extrato = st.file_uploader(
            "Upload extrato do Asaas", 
            type=['xlsx'],
            key="extrato_asaas",
            help="Extrato de movimentações da conta Asaas"
        )
    
    # Botão para usar arquivos padrão
    if st.button("Usar Arquivos Padrão do Sistema"):
        try:
            import os
            arquivo_cobrancas_padrao = "Cobranc_as (5).xlsx"
            arquivo_extrato_padrao = "Extrato_Asaas.xlsx"
            
            if os.path.exists(arquivo_cobrancas_padrao) and os.path.exists(arquivo_extrato_padrao):
                if contas_asaas.carregar_dados(arquivo_cobrancas_padrao, arquivo_extrato_padrao):
                    st.success("Arquivos padrão carregados com sucesso!")
                    st.session_state.asaas_dados_carregados = True
                else:
                    st.error("Erro ao carregar arquivos padrão")
            else:
                st.error("Arquivos padrão não encontrados no sistema")
        except Exception as e:
            st.error(f"Erro: {e}")
    
    # Processar uploads
    if arquivo_cobrancas and arquivo_extrato:
        if contas_asaas.carregar_dados(arquivo_cobrancas, arquivo_extrato):
            st.success("Arquivos carregados com sucesso!")
            st.session_state.asaas_dados_carregados = True
        else:
            st.error("Erro ao processar arquivos")
    
    # Mostrar análise se dados estão carregados
    if st.session_state.get('asaas_dados_carregados', False) or (arquivo_cobrancas and arquivo_extrato):
        
        st.divider()
        
        # Filtros de data
        st.subheader("Filtros de Período")
        col1, col2, col3 = st.columns([1, 1, 1])
        
        with col1:
            data_inicio = st.date_input(
                "Data Início",
                value=datetime(2025, 7, 1),
                help="Filtrar dados a partir desta data"
            )
            
        with col2:
            data_fim = st.date_input(
                "Data Fim", 
                value=datetime.now(),
                help="Filtrar dados até esta data"
            )
            
        with col3:
            st.markdown("**Ações**")
            atualizar = st.button("Atualizar Análise", type="primary")
        
        # Obter métricas
        try:
            # Primeiro carregar métricas gerais para mostrar período dos dados
            metricas_completas = contas_asaas.obter_metricas_dashboard('completo')
            
            # Mostrar período dos dados disponíveis
            if 'periodo_dados' in metricas_completas:
                st.info(f"""
                **Período dos dados disponíveis:**
                - Vendas: {metricas_completas['periodo_dados']['vendas_min'].strftime('%d/%m/%Y')} até {metricas_completas['periodo_dados']['vendas_max'].strftime('%d/%m/%Y')}
                - Extrato: {metricas_completas['periodo_dados']['extrato_min'].strftime('%d/%m/%Y')} até {metricas_completas['periodo_dados']['extrato_max'].strftime('%d/%m/%Y')}
                """)
            
            # Obter reconciliação para o período filtrado
            reconciliacao = contas_asaas.reconciliar_vendas_recebimentos(data_inicio, data_fim)
            
            # Cards de métricas principais
            st.subheader("Métricas Consolidadas")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "Vendas (Competência)",
                    f"R$ {reconciliacao['resumo']['total_vendas_competencia']:,.2f}",
                    help="Total de vendas registradas no período (competência)"
                )
                
            with col2:
                st.metric(
                    "Recebido (Caixa)",
                    f"R$ {reconciliacao['resumo']['total_recebido_caixa']:,.2f}",
                    help="Valores efetivamente recebidos no caixa"
                )
                
            with col3:
                st.metric(
                    "Antecipado",
                    f"R$ {reconciliacao['resumo']['total_antecipado']:,.2f}",
                    help="Valores antecipados de recebíveis"
                )
                
            with col4:
                saldo_a_receber = reconciliacao['resumo']['saldo_a_receber']
                delta_color = "normal" if saldo_a_receber >= 0 else "inverse"
                st.metric(
                    "Saldo a Receber",
                    f"R$ {saldo_a_receber:,.2f}",
                    help="Vendas ainda não recebidas nem antecipadas"
                )
            
            # Card adicional com taxas
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(
                    "Taxas Antecipação",
                    f"R$ {reconciliacao['resumo']['total_taxas_antecipacao']:,.2f}",
                    help="Taxas pagas pelas antecipações"
                )
                
            # Gráficos comparativos
            st.subheader("Análise Gráfica")
            
            if not reconciliacao['vendas_detalhadas'].empty and not reconciliacao['recebimentos_detalhados'].empty:
                
                # Gráfico de vendas vs recebimentos por data
                fig = make_subplots(
                    rows=2, cols=1,
                    subplot_titles=('Vendas por Data (Competência)', 'Recebimentos por Data (Caixa)'),
                    vertical_spacing=0.1
                )
                
                # Vendas
                vendas = reconciliacao['vendas_detalhadas']
                fig.add_trace(
                    go.Bar(
                        x=vendas['Data_Venda'],
                        y=vendas['Valor_Bruto'],
                        name="Vendas",
                        marker_color='#1e3c72'
                    ),
                    row=1, col=1
                )
                
                # Recebimentos
                recebimentos = reconciliacao['recebimentos_detalhadas']
                fig.add_trace(
                    go.Bar(
                        x=recebimentos['Data_Movimento'],
                        y=recebimentos['Valor_Recebido'],
                        name="Recebimentos",
                        marker_color='#2a5298'
                    ),
                    row=2, col=1
                )
                
                fig.update_layout(
                    height=600,
                    title_text="Vendas vs Recebimentos - Análise Temporal",
                    showlegend=True
                )
                
                st.plotly_chart(fig, use_container_width=True)
            
            # Tabelas detalhadas
            st.subheader("Detalhamento")
            
            tab1, tab2, tab3 = st.tabs(["Vendas", "Recebimentos", "Antecipações"])
            
            with tab1:
                if not reconciliacao['vendas_detalhadas'].empty:
                    st.markdown("**Vendas por Data (Competência)**")
                    df_vendas_show = reconciliacao['vendas_detalhadas'].copy()
                    df_vendas_show['Valor_Bruto'] = df_vendas_show['Valor_Bruto'].apply(lambda x: f"R$ {x:,.2f}")
                    df_vendas_show['Valor_Liquido'] = df_vendas_show['Valor_Liquido'].apply(lambda x: f"R$ {x:,.2f}")
                    st.dataframe(df_vendas_show, use_container_width=True)
                else:
                    st.info("Nenhuma venda encontrada no período")
            
            with tab2:
                if not reconciliacao['recebimentos_detalhadas'].empty:
                    st.markdown("**Recebimentos por Data (Caixa)**")
                    df_receb_show = reconciliacao['recebimentos_detalhadas'].copy()
                    df_receb_show['Valor_Recebido'] = df_receb_show['Valor_Recebido'].apply(lambda x: f"R$ {x:,.2f}")
                    st.dataframe(df_receb_show, use_container_width=True)
                else:
                    st.info("Nenhum recebimento encontrado no período")
            
        
        except Exception as e:
            st.error(f"❌ Erro ao carregar dados do Asaas: {e}")


def main():
    """Função principal do dashboard"""
    # Verifica autenticação
    check_authentication()
    
    # Mostra informações do usuário
    show_user_info()
    
    # Botão atualizar
    if st.sidebar.button("Atualizar Dados", type="primary"):
        st.cache_data.clear()
        st.rerun()
    
    # Navegação
    secao = st.sidebar.radio(
        "Navegação:",
        [
            "Realizado",
            "Gráficos",
            "Fluxo de Caixa",
            "Faturamento Tempo Real",
            "Projeções"
        ]
    )
    
    # Carrega dados
    df_despesas, df_receitas, df_fluxo = carregar_dados()
    
    if df_despesas is None:
        st.error("❌ Erro ao carregar dados. Execute primeiro o processamento corrigido.")
        st.code("python processar_despesas_correto.py")
        return
    
    # Roteamento
    if secao == "Realizado":
        # Combine overview cards + matriz realizada em uma única seção
        # Overview cards integrado na matriz realizada
        st.divider()
        mostrar_matriz_realizada(df_despesas, df_receitas)
    
    elif secao == "Gráficos":
        mostrar_graficos_avancados(df_despesas, df_receitas)
    
    elif secao == "Fluxo de Caixa":
        mostrar_fluxo_caixa_projecoes()
    
    elif secao == "Faturamento Tempo Real":
        from modulo_faturamento_tempo_real import main_modulo_faturamento_tempo_real
        main_modulo_faturamento_tempo_real()
    
    elif secao == "Projeções":
        mostrar_projecoes_5_anos()

if __name__ == "__main__":
    main()
